from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file,send_from_directory
from datetime import datetime,timedelta
import os
import io
import tempfile
import copy
from solver import optimize_crude_mix_schedule
from collections import defaultdict
import re
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import the Simulator class from your scheduler script
from scheduler import Simulator
SUMMARY_FILE_PATH = "/tmp/daily_summary.csv"

# --- User Authentication ---
APP_USERNAME = os.environ.get("APP_USERNAME", "admin")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "admin123")


# --- Helper Functions for Parameter Validation ---
def safe_float(value, default=0.0, param_name="unknown"):
    """Safely convert value to float with logging"""
    try:
        if value is None or value == '':
            # print(f"WARNING: {param_name} is None or empty, using default {default}")
            return default
        result = float(value)
        # print(f"DEBUG: {param_name} = {result}")
        return result
    except (ValueError, TypeError) as e:
        print(f"ERROR: Cannot convert {param_name}={value} to float: {e}")
        raise ValueError(f"Invalid value for {param_name}: {value}")


def safe_int(value, default=0, param_name="unknown"):
    """Safely convert value to int with logging"""
    try:
        if value is None or value == '':
            # print(f"WARNING: {param_name} is None or empty, using default {default}")
            return default
        result = int(value)
        # print(f"DEBUG: {param_name} = {result}")
        return result
    except (ValueError, TypeError) as e:
        print(f"ERROR: Cannot convert {param_name}={value} to int: {e}")
        raise ValueError(f"Invalid value for {param_name}: {value}")


def _parse_json_datetime(dt_str):
    """Parse datetime string from JSON"""
    if not dt_str:
        return None
    try:
        return datetime.strptime(dt_str, '%d/%m/%Y %H:%M')
    except:
        try:
            return datetime.strptime(dt_str, '%d/%m/%y %H:%M')
        except:
            return None

def _parse_log_dt(dt_val):
    """Safely parse a datetime object or a string from various formats."""
    if not dt_val: return None
    if isinstance(dt_val, datetime): return dt_val
    if isinstance(dt_val, str):
        for fmt in ['%d/%m/%Y %H:%M', '%d/%m/%y %H:%M', '%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%d %H:%M:%S']:
            try: return datetime.strptime(dt_val, fmt)
            except (ValueError, TypeError): continue
    return None

def _parse_tank_id_from_str(tank_str):
    """Extracts the integer ID from a string like 'Tank14'."""
    if not tank_str: return None
    match = re.search(r'\d+', tank_str)
    if match: return int(match.group(0))
    return None

def _build_cycle_data_from_log(simulation_log):
    """Processes simulation_log to structure event timestamps by tank and cycle."""
    # Structure: cycles_data[tank_id][cycle_num] = {event_key: timestamp, ...}
    cycles_data = defaultdict(lambda: defaultdict(dict))
    # Track the latest timestamp seen for each event type within a cycle to avoid overwrites by earlier logs
    latest_event_time = defaultdict(lambda: defaultdict(lambda: datetime.min))


    # Sort log chronologically first to process events in order
    sorted_log = sorted(
        [entry for entry in simulation_log if isinstance(entry, dict)], # Filter out non-dict entries
        key=lambda x: _parse_log_dt(x.get("Timestamp")) or datetime.min
    )

    for entry in sorted_log:
        event_str = entry.get("Event", "")
        timestamp = _parse_log_dt(entry.get("Timestamp"))
        tank_id = _parse_tank_id_from_str(entry.get("Tank"))

        if not (event_str and timestamp and tank_id):
            continue # Skip if essential info is missing

        # Extract base event name and cycle number (e.g., "READY", 1 from "READY_1")
        match = re.match(r'([A-Z_]+)_(\d+)$', event_str)
        cycle_num = None
        base_event = event_str # Default if no cycle number found
        if match:
            base_event = match.group(1)
            cycle_num = int(match.group(2))
        else:
             # Skip events that aren't cycle-specific if we only care about cycles
             if base_event not in ["FILL_START_FIRST"]: # Allow initial fill without number? Decide based on log. Assuming starts need _1.
                 continue # Ignore non-cycle events like SIM_START, ARRIVAL etc.

        # Map the base log event names to the keys we want in our cycle data dict
        event_map = {
            "FILL_START_FIRST": "fill_start",
            # Add "FILL_START" if subsequent fills start a new cycle in your logic
            "FILL_FINAL_END": "fill_end",
            "SETTLING_START": "settle_start",
            "SETTLING_END": "settle_end", # We use this to infer lab_start
            "READY": "ready"
        }

        if base_event in event_map:
            key = event_map[base_event]
            cycle_key = (tank_id, cycle_num)

            # --- Store timestamp only if it's the latest seen for this event in this cycle ---
            # This handles potential duplicate log entries or out-of-order processing within the same timestamp
            if timestamp >= latest_event_time[cycle_key][key]:
                 cycles_data[tank_id][cycle_num][key] = timestamp
                 latest_event_time[cycle_key][key] = timestamp


    # --- Infer lab_start from settle_end ---
    # This assumes lab testing starts immediately after settling ends.
    # If your simulation has a different logic (e.g., a specific LAB_START event), adjust this.
    for tank_id, cycles in cycles_data.items():
        for cycle_num, data in cycles.items():
            if 'settle_end' in data and 'lab_start' not in data:
                # Add lab_start timestamp, assuming it's the same as settle_end
                cycles_data[tank_id][cycle_num]['lab_start'] = data['settle_end']

    return cycles_data
def _parse_detail_start_time(detail_str):
    """Parses the start datetime from 'dd/mm/yyyy hh:mm-...'"""
    if not detail_str: return None
    try:
        # Match dd/mm/yyyy hh:mm at the beginning
        match = re.match(r'(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})', detail_str)
        if match:
            return datetime.strptime(match.group(1), '%d/%m/%Y %H:%M')
    except (ValueError, TypeError):
        pass
    return None # Return None if parsing fails
def _parse_sheet_datetime(date_str, time_str):
    """Parses date and time strings (e.g., from cargo_report) into a datetime object."""
    if not date_str or not time_str:
        return None
    try:
        # Assuming format 'dd/mm/yyyy hh:mm' based on cargo_report examples
        return datetime.strptime(f"{date_str} {time_str}", '%d/%m/%Y %H:%M')
    except (ValueError, TypeError):
        # Add fallback formats if needed
        try:
             # Try dd/mm/yy format as a fallback
             return datetime.strptime(f"{date_str} {time_str}", '%d/%m/%y %H:%M')
        except (ValueError, TypeError):
             print(f"Warning: Could not parse date/time from sheet: '{date_str} {time_str}'")
             return None # Return None if parsing fails

def _format_timedelta(delta):
    """Formats a timedelta into Xxd Yyh ZZm"""
    if not isinstance(delta, timedelta) or delta.total_seconds() < 0:
        return "N/A" # Return N/A for invalid input

    total_seconds = int(delta.total_seconds())
    days = total_seconds // (24 * 3600)
    seconds_remaining = total_seconds % (24 * 3600)
    hours = seconds_remaining // 3600
    seconds_remaining %= 3600
    minutes = seconds_remaining // 60

    return f"{days:02d}d {hours:02d}h {minutes:02d}m"

# --- Sheet Creation Helper Functions ---
def _create_simulation_log_sheet(wb, results):
    """Create Sheet 1: Simulation Log with all events"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import re
        
        ws = wb.create_sheet("simulation_log")
        
        daily_log = results.get('simulation_log', [])
        if not daily_log:
            return False
        
        timestamp_str = datetime.now().strftime('%d/%m/%y %H:%M')
        ws.cell(row=1, column=1, value=timestamp_str)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        fixed_columns = ['Timestamp', 'Level', 'Event', 'Tank', 'Cargo', 'Message']
        
        tank_columns = sorted([
            key for key in daily_log[0].keys() 
            if re.match(r'^Tank\d+$', key)
        ], key=lambda x: int(re.search(r'\d+', x).group(0)))
        
        headers = fixed_columns + tank_columns
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, row_data in enumerate(daily_log, 3):
            event = row_data.get('Event', '')
            ready_count = 0
            
            if event.startswith('READY'):
                log_index = row_idx - 3
                if log_index + 1 < len(daily_log):
                    next_row = daily_log[log_index + 1]
                    ready_count = sum(1 for tank_col in tank_columns if next_row.get(tank_col) == 'READY')
                else:
                    ready_count = sum(1 for tank_col in tank_columns if row_data.get(tank_col) == 'READY')
            
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                
                if header == 'Message' and event.startswith('READY') and ready_count > 0:
                    value = f"{value} No of READY tanks : {ready_count}"
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if header == 'Level':
                    if value.lower() == 'danger':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif value.lower() == 'warning':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif value.lower() == 'success':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
                if re.match(r'^Tank\d+$', header):
                    if value == 'READY':
                        cell.fill = PatternFill(start_color='006100', end_color='006100', fill_type='solid')
                    elif value == 'EMPTY':
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    elif value == 'SUSPENDED':
                        cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                    elif value == 'LAB':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif value == 'SETTLING':
                        cell.fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
                    elif value == 'FEEDING':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif value in ['FILLING', 'FILLED']:
                        cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                
                if header in ['Timestamp', 'Message', 'Cargo']:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True if header == 'Message' else False)
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            header_name = headers[col_idx - 1]
            
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            if header_name == 'Message':
                adjusted_width = min(max(max_length + 5, 50), 100)
            else:
                adjusted_width = min(max(max_length + 3, 12), 40)
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A3'
        
        return True
        
    except Exception as e:
        print(f"Error creating simulation log sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def _create_daily_summary_sheet(wb, results):
    """Create Sheet 2: Daily summary with enhanced tank status columns"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import re
        
        ws = wb.create_sheet("daily_summary_sheet")
        
        simulation_data = results.get('simulation_data', [])
        simulation_log = results.get('simulation_log', [])
        parameters = results.get('parameters', {})
        
        if not simulation_data:
            return False
        
        timestamp_str = datetime.now().strftime('%d/%m/%y %H:%M')
        ws.cell(row=1, column=1, value=timestamp_str)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:N1') # <-- MODIFIED MERGE
        
        # --- START MODIFICATION ---
        base_headers = ['DAY', 'Date', 'Opening Stock (bbl)', 'Certified (bbl)', 'Uncertified (bbl)',
                       'Processing (bbl)', 'Closing Stock (bbl)', 'Tank Util %',
                       'Ready Tanks', 'Empty Tanks', 'FILLING', 'LAB', 'FEEDING', 
                       'SUSPENDED', 'TOTAL']
        # --- END MODIFICATION ---
        
        num_tanks = sum(1 for key in simulation_data[0].keys() 
                      if key.startswith('Tank') and len(key) <= 6)
        tank_headers = [f'Tank{i}' for i in range(1, num_tanks + 1)]
        headers = base_headers + tank_headers
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, row_data in enumerate(simulation_data, 3):
            day_number = row_idx - 2
            date_str = row_data.get('date', '')
            opening_stock = row_data.get('start_inventory', 0)
            closing_stock = row_data.get('end_inventory', 0)
            processing = row_data.get('processing', 0)
            certified_stock = row_data.get('certified_stock', 0)
            tank_utilization = row_data.get('tank_utilization', 0)
            
            uncertified_stock = max(0, opening_stock - certified_stock)
            
            # Find DAILY_STATUS event from simulation_log for this date to get tank statuses
            current_date = date_str.split(' ')[0] if date_str else ""
            tank_statuses_from_log = {}
            
            if simulation_log and current_date:
                daily_status_log = None
                for log in simulation_log:
                    if log.get('Event') == 'DAILY_STATUS':
                        log_date = log.get('Timestamp', '').split(' ')[0]
                        if log_date == current_date:
                            daily_status_log = log
                            break
                
                if daily_status_log:
                    for tank_num in range(1, num_tanks + 1):
                        tank_key = f'Tank{tank_num}'
                        tank_statuses_from_log[tank_key] = daily_status_log.get(tank_key, '')
            
            # --- START MODIFICATION ---
            # Count tank statuses from simulation_log
            ready_tanks_count = sum(1 for t in tank_statuses_from_log.values() if t == 'READY')
            empty_tanks_count = sum(1 for t in tank_statuses_from_log.values() if t == 'EMPTY')
            filling_count = sum(1 for t in tank_statuses_from_log.values() if t in ['FILLING', 'FILLED'])
            lab_count = sum(1 for t in tank_statuses_from_log.values() if t in ['LAB', 'SETTLING'])
            feeding_count = sum(1 for t in tank_statuses_from_log.values() if t == 'FEEDING')
            suspended_count = sum(1 for t in tank_statuses_from_log.values() if t == 'SUSPENDED') # <-- ADDED COUNT
            total_count = (ready_tanks_count + empty_tanks_count + filling_count + 
                           lab_count + feeding_count + suspended_count) # <-- UPDATED TOTAL
            # --- END MODIFICATION ---
            
            col_idx = 1
            for header in base_headers:
                if header == 'DAY':
                    value = day_number
                elif header == 'Date':
                    value = date_str
                elif header == 'Opening Stock (bbl)':
                    value = opening_stock
                elif header == 'Certified (bbl)':
                    value = certified_stock
                elif header == 'Uncertified (bbl)':
                    value = uncertified_stock
                elif header == 'Processing (bbl)':
                    value = processing
                elif header == 'Closing Stock (bbl)':
                    value = closing_stock
                elif header == 'Tank Util %':
                    value = tank_utilization
                elif header == 'Ready Tanks':
                    value = ready_tanks_count
                elif header == 'Empty Tanks':
                    value = empty_tanks_count
                elif header == 'FILLING':
                    value = filling_count
                elif header == 'LAB':
                    value = lab_count
                elif header == 'FEEDING':
                    value = feeding_count
                # --- START MODIFICATION ---
                elif header == 'SUSPENDED':
                    value = suspended_count
                # --- END MODIFICATION ---
                elif header == 'TOTAL':
                    value = total_count
                else:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if header == 'Date':
                    cell.alignment = Alignment(horizontal='left')
                elif header == 'Tank Util %':
                    cell.alignment = Alignment(horizontal='right')
                    if value:
                        cell.number_format = '0.0"%"'
                elif isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    if value >= 1000:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                col_idx += 1
            
            for tank_num in range(1, num_tanks + 1):
                tank_key = f'Tank{tank_num}'
                tank_state = tank_statuses_from_log.get(tank_key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=tank_state)
                cell.alignment = Alignment(horizontal='center')
                
                # --- START COLOR CHANGE ---
                if tank_state == 'READY':
                    # Light Green
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif tank_state == 'EMPTY':
                    # Pure Red
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif tank_state == 'SUSPENDED':
                    # Light Red
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif tank_state == 'FEEDING':
                    # Yellow
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif tank_state in ['FILLING', 'FILLED']:
                    # Light Blue
                    cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                elif tank_state in ['LAB', 'SETTLING']:
                    # Light Olive
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                # --- END COLOR CHANGE ---
                
                col_idx += 1
        
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in ws[column_letter]:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 3, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return True
        
    except Exception as e:
        print(f"Error creating daily summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return False
        
def _create_cargo_arrivals_sheet(wb, results):
    """
    Create Sheet 3: Cargo arrivals timeline with PRE DISCHARGE duration
    and tank fill columns showing volume and crude type per tank.
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import re
        
        ws = wb.create_sheet("cargo_arrival_sheet")

        cargo_report = results.get('cargo_report', [])
        simulation_log = results.get('simulation_log', [])
        parameters = results.get('parameters', {})

        if not cargo_report:
            ws.cell(row=1, column=1, value="No cargo report data available.")
            return False

        # Build tank_mix from simulation_log FILL_START events
        tank_crude_map = {}  # {(vessel_name, tank_id): crude_type}
        for log in simulation_log:
            event = log.get('Event', '')
            if event.startswith('FILL_START'):
                message = log.get('Message', '')
                tank = log.get('Tank', '')
                cargo = log.get('Cargo', '')
                
                # Extract tank ID from "Tank 1" format
                tank_match = re.search(r'Tank (\d+)', tank)
                if tank_match and cargo:
                    tank_id = int(tank_match.group(1))
                    
                    # Extract crude from message like "filling Tank 1 with 589,750 bbl Basrah Light"
                    crude_match = re.search(r'bbl\s+(.+?)\s*\(', message)
                    if not crude_match:
                        crude_match = re.search(r'bbl\s+(.+?)$', message)
                    
                    if crude_match:
                        crude_type = crude_match.group(1).strip()
                        tank_crude_map[(cargo, tank_id)] = crude_type

        # --- Parse tank fills with volume and crude info ---
        earliest_tank_fill_time = {}
        parsed_cargo_fills_list = []
        
        for cargo_data in cargo_report:
            tank_fill_details_str = cargo_data.get('Tank Fill Details', '')
            vessel_name = cargo_data.get('Vessel Name', '')
            current_cargo_parsed_fills = {}
            
            if tank_fill_details_str:
                for entry in tank_fill_details_str.split(' | '):
                    # Match: "Tank1: Tank1: 14/08/2025 09:37-14/08/2025 18:49 (589,750 bbl)"
                    match = re.match(r'(Tank(\d+)):\s*Tank\d+:\s*(.+?)\s*\(([0-9,]+)\s*bbl\)', entry.strip())
                    if not match:
                        match = re.match(r'(Tank(\d+)):\s*(.+?)\s*\(([0-9,]+)\s*bbl\)', entry.strip())
                    
                    if match:
                        tank_name, tank_id_str, time_range, volume_str = match.groups()
                        tank_id = int(tank_id_str)
                        
                        # Get crude type for this specific tank fill
                        crude_type = tank_crude_map.get((vessel_name, tank_id), 'Unknown')
                        
                        # Format: "589,750 bbl - Crude Name\n14/08/2025 09:37-14/08/2025 18:49"
                        formatted_detail = f"{volume_str} bbl - {crude_type}\n{time_range.strip()}"
                        current_cargo_parsed_fills[tank_name] = formatted_detail
                        
                        start_time = _parse_detail_start_time(time_range)
                        if start_time:
                            if tank_id not in earliest_tank_fill_time or start_time < earliest_tank_fill_time[tank_id]:
                                earliest_tank_fill_time[tank_id] = start_time
            
            parsed_cargo_fills_list.append(current_cargo_parsed_fills)

        # Sort Tank IDs based on earliest fill time
        all_filled_tank_ids = list(earliest_tank_fill_time.keys())
        sorted_tank_ids = sorted(all_filled_tank_ids, key=lambda tid: earliest_tank_fill_time.get(tid, datetime.max))

        # Add timestamp
        timestamp_str = datetime.now().strftime('%d/%m/%y %H:%M')
        ws.cell(row=1, column=1, value=timestamp_str).font = Font(bold=True, size=12)

        # --- Define Headers with PRE DISCHARGE ---
        base_headers = [
            'Vessel Name', 'Berth',
            'Arrival Date', 'Arrival Time',
            'Discharge Start Date', 'Discharge Start Time',
            'PRE DISCHARGE',
            'Discharge End Date', 'Discharge End Time',
            'BERTH GAP (hrs)',
            'Discharge Duration (hours)', 'Total Volume Discharged (bbl)',
            'Tanks Filled'
        ]
        tank_fill_headers = [f'Tank{tank_id} Fill' for tank_id in sorted_tank_ids]
        headers = base_headers + tank_fill_headers

        # Adjust merge cells
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # Write headers (row 2)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data rows (starting from row 3)
        for row_idx, (cargo_data, parsed_fills) in enumerate(zip(cargo_report, parsed_cargo_fills_list), 3):

            # --- Parse Arrival and Fill Start Times ---
            arrival_date_str = cargo_data.get('Arrival Date', '')
            arrival_time_str = cargo_data.get('Arrival Time', '')
            fill_start_date_str = cargo_data.get('Discharge Start Date', '')
            fill_start_time_str = cargo_data.get('Discharge Start Time', '')

            arrival_dt = _parse_sheet_datetime(arrival_date_str, arrival_time_str)
            fill_start_dt = _parse_sheet_datetime(fill_start_date_str, fill_start_time_str)

            # --- Calculate PRE DISCHARGE Duration ---
            pre_discharge_duration_str = "N/A"
            if arrival_dt and fill_start_dt and fill_start_dt >= arrival_dt:
                duration_delta = fill_start_dt - arrival_dt
                pre_discharge_duration_str = _format_timedelta(duration_delta)

            # --- Write Base Data ---
            col_idx = 1
            for header in base_headers:
                if header == 'PRE DISCHARGE':
                    value = pre_discharge_duration_str
                else:
                    value = cargo_data.get(header, '')

                # Clean numeric values if they are strings
                if isinstance(value, str):
                    if header == 'Total Volume Discharged (bbl)':
                        try:
                            value = float(value.replace(',', '').replace(' bbl', ''))
                        except ValueError:
                            value = 0.0
                    elif header in ['Discharge Duration (hours)', 'Tanks Filled', 'BERTH GAP (hrs)']:
                        if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                            try:
                                value = float(value)
                            except ValueError:
                                pass

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Formatting
                if header == 'BERTH GAP (hrs)':
                    cell.alignment = Alignment(horizontal='center')
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                elif isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    if header == 'Total Volume Discharged (bbl)':
                        cell.number_format = '#,##0'
                    elif header in ['Discharge Duration (hours)', 'Tanks Filled']:
                        cell.number_format = '0.00'
                elif header.endswith('Date') or header.endswith('Time') or header == 'PRE DISCHARGE':
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

                if header == 'Berth':
                    if 'BERTH 1' in str(value):
                        cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
                    elif 'BERTH 2' in str(value):
                        cell.fill = PatternFill(start_color='FCE4D6', fill_type='solid')

                col_idx += 1

            # --- Write tank fill details with volume and crude ---
            for tank_id in sorted_tank_ids:
                tank_key = f'Tank{tank_id}'
                tank_fill_detail = parsed_fills.get(tank_key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=tank_fill_detail)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                col_idx += 1

        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            header_name = headers[col_idx - 1]

            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Adjust width based on column type
            if header_name.endswith(' Fill'):
                adjusted_width = min(max(max_length + 2, 30), 60)
            elif header_name == 'Vessel Name':
                adjusted_width = min(max(max_length + 2, 18), 30)
            elif header_name == 'Berth':
                adjusted_width = 10
            elif header_name.endswith('(hours)'):
                adjusted_width = 15
            elif header_name.endswith('(bbl)'):
                adjusted_width = 20
            elif header_name.endswith('Date') or header_name.endswith('Time'):
                adjusted_width = 18
            elif header_name == 'PRE DISCHARGE':
                adjusted_width = 18
            elif header_name == 'BERTH GAP (hrs)':
                adjusted_width = 18
            else:
                adjusted_width = min(max(max_length + 2, 12), 40)

            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header rows and first column
        ws.freeze_panes = 'B3'

        return True

    except Exception as e:
        print(f"Error creating cargo arrivals sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def _create_certified_stock_chart_sheet(wb, results):
    """
    Create a sheet with data and an embedded chart for Certified Stock vs. Day.
    """
    try:
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.chart.axis import TextAxis, NumericAxis
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
        
        ws = wb.create_sheet("Certified_Stock_Chart")
        
        simulation_data = results.get('simulation_data', [])
        if not simulation_data:
            ws.cell(row=1, column=1, value="No simulation data available for charting.")
            return False

        # 1. Write Headers (Row 1) - ONLY Certified Stock
        headers = ['Day', 'Certified Stock (bbl)']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2. Write Data (Starting Row 2)
        for row_idx, row_data in enumerate(simulation_data, 2):
            day = row_idx - 1
            
            # Use 'certified_stock' from the transformed data
            certified_stock = row_data.get('certified_stock', 0)
            
            ws.cell(row=row_idx, column=1, value=day).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=2, value=certified_stock).number_format = '#,##0'
            
        max_row = len(simulation_data) + 1
        
        # 3. Create Chart with proper sizing
        chart = LineChart()
        chart.title = "Certified Crude Oil Stock vs. Day"
        
        # Set chart dimensions (width and height in cm)
        chart.width = 20  # Width in cm (default is ~15)
        chart.height = 12  # Height in cm (default is ~7.5)
        
        # Configure X-axis explicitly
        chart.x_axis.title = "Day of Simulation"
        chart.x_axis.tickLblPos = "low"  # Position tick labels below axis
        
        # Configure Y-axis explicitly  
        chart.y_axis.title = "Certified Stock (bbl)"
        chart.y_axis.tickLblPos = "low"  # Position tick labels to the left
        
        # Define x-axis labels (Day numbers - Column A)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row) 
        chart.set_categories(labels_ref)
        
        # Define data series for Certified Stock
        certified_stock_data = Reference(ws, min_col=2, min_row=1, max_row=max_row)  # Include header
        series_certified = Series(certified_stock_data, title_from_data=True)
        
        # Style the line properly
        line_props = LineProperties()
        line_props.solidFill = '0066CC'  # Blue color
        line_props.width = 25000  # Line width in EMUs (English Metric Units)
        series_certified.graphicalProperties.line = line_props
        
        chart.series.append(series_certified)
        
        # Add legend
        chart.legend.position = 'r'  # Position legend on the right
        
        # Position the chart below the data with more space
        ws.add_chart(chart, "E2")
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25

        return True
        
    except Exception as e:
        print(f"Error creating certified stock chart sheet: {e}")
        import traceback
        traceback.print_exc()
        return False
def _create_sequence_summary_sheets(wb, results):
    """Creates the Sequence Summary sheet using cycle data extracted from simulation_log."""
    try:
        ws = wb.create_sheet("Sequence Summary")

        timestamp_str = f"Report Generated On: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
        ws.cell(row=1, column=1, value=timestamp_str).font = Font(bold=True, italic=True, color="4F4F4F")

        tank_details = results.get('full_tank_details', [])
        tank_name_map = {tank.get('id'): (tank.get('display_name') or f"Tank {tank.get('id')}") for tank in tank_details if tank.get('id')}

        cargo_report = results.get('cargo_report', []) or []
        feeding_events_log = results.get('feeding_events_log', []) or []
        simulation_log = results.get('simulation_log', []) or [] # Get the simulation log

        # --- START FIX: Extract mix from the READY log, not from daily_discharge_log ---
        tank_mix_map = {} # This will store { (tank_id, cycle_num): "Mix String" }
        
        for entry in simulation_log:
            event_str = entry.get("Event", "")
            
            # Check if this is a READY event (e.g., "READY_1", "READY_2")
            if event_str.startswith("READY_"):
                message = entry.get("Message", "")
                tank_id = _parse_tank_id_from_str(entry.get("Tank"))
                
                if not tank_id:
                    continue
                    
                # Extract cycle number from "READY_X"
                try:
                    cycle_num = int(event_str.split('_')[-1])
                except (ValueError, IndexError):
                    continue

                # Extract the mix string from the message
                # Message format: "Tank 14 now READY - Mix: [Bonny Light: 50.0%, ...]"
                mix_match = re.search(r'Mix:\s*\[(.*?)\]', message)
                if mix_match:
                    mix_string = mix_match.group(1).strip()
                    # Store the mix string against the tank and its specific cycle
                    tank_mix_map[(tank_id, cycle_num)] = mix_string
        # --- END FIX ---

        # --- Call the function to build structured cycle data ---
        cycles_data = _build_cycle_data_from_log(simulation_log)

        current_row = 3

        # --- CARGO SEQUENCE Table (Remains the same) ---
        ws.cell(row=current_row, column=1, value="CARGO SEQUENCE").font = Font(bold=True, size=14)
        current_row += 2
        cargo_headers = ['CARGO', 'ARRIVAL_DATE', 'ARRIVAL_TIME', 'DEPARTURE_DATE', 'DEPARTURE_TIME']
        for col, header in enumerate(cargo_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        for cargo in cargo_report:
             row_data = [ cargo.get('Cargo Type', '').title(), cargo.get('Arrival Date', ''), cargo.get('Arrival Time', ''), cargo.get('Discharge End Date', ''), cargo.get('Discharge End Time', '') ]
             for col, value in enumerate(row_data, 1): ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
             current_row += 1
        current_row += 2

        # --- FEEDING SEQUENCE Table (Remains the same) ---
        ws.cell(row=current_row, column=1, value="FEEDING SEQUENCE").font = Font(bold=True, size=14)
        current_row += 2
        feeding_headers = ['TANK', 'START_DATE', 'START_TIME', 'END_DATE', 'END_TIME']
        for col, header in enumerate(feeding_headers, 1):
             cell = ws.cell(row=current_row, column=col, value=header)
             cell.font = Font(bold=True, color="FFFFFF")
             cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
             cell.alignment = Alignment(horizontal='center')
        current_row += 1
        sorted_feeding_log = sorted(feeding_events_log, key=lambda x: _parse_log_dt(x.get('start')) or datetime.min)
        for event in sorted_feeding_log:
             start_dt = _parse_log_dt(event.get('start'))
             end_dt = _parse_log_dt(event.get('end'))
             tank_id = event.get('tank_id')
             row_data = [ tank_name_map.get(tank_id, f"Tank {tank_id}"), start_dt.strftime('%d/%m/%y') if start_dt else '', start_dt.strftime('%H:%M') if start_dt else '', end_dt.strftime('%d/%m/%y') if end_dt else 'N/A', end_dt.strftime('%H:%M') if end_dt else 'N/A' ]
             for col, value in enumerate(row_data, 1): ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
             current_row += 1
        current_row += 2

        # --- FILLING, SETTLING & LAB TESTING SEQUENCE Table (Uses new cycles_data with chronological sort) ---
        ws.cell(row=current_row, column=1, value="FILLING, SETTLING & LAB TESTING SEQUENCE").font = Font(bold=True, size=14)
        current_row += 2
        
        processing_headers = [
            'TANK','TANK MIX',
            'FILL_START_DATE', 'FILL_START_TIME', 
            'FILL_END_DATE', 'FILL_END_TIME', 
            'SETTLE_START_DATE', 'SETTLE_START_TIME', 
            'LABTEST_START_DATE', 'LABTEST_START_TIME', 
            'READY_DATE', 'READY_TIME',
            'TANK READY TIME'
        ]
        
        for col, header in enumerate(processing_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # --- Create a flat list of all cycles ---
        all_cycles = []
        for tank_id, cycles in cycles_data.items():
            for cycle_num, cycle_timestamps in cycles.items():
                fill_start = cycle_timestamps.get('fill_start')
                if fill_start: # Only include cycles with a start time
                    all_cycles.append({
                        'tank_id': tank_id,
                        'cycle_num': cycle_num,
                        'timestamps': cycle_timestamps,
                        'fill_start': fill_start # For sorting
                    })

        # --- Sort the flat list chronologically by fill_start time ---
        sorted_all_cycles = sorted(all_cycles, key=lambda x: x['fill_start'])

        # --- Iterate through the chronologically sorted cycles ---
        for cycle_info in sorted_all_cycles:
            tank_id = cycle_info['tank_id']
            cycle_num = cycle_info['cycle_num'] 
            cycle_timestamps = cycle_info['timestamps']

            fill_start   = cycle_timestamps.get('fill_start') # Known to exist
            fill_end     = cycle_timestamps.get('fill_end')
            settle_start = cycle_timestamps.get('settle_start')
            lab_start    = cycle_timestamps.get('lab_start') # Inferred
            ready_time   = cycle_timestamps.get('ready')

            # Calculate the duration
            tank_ready_time_str = "N/A"
            if fill_start and ready_time:
                duration_delta = ready_time - fill_start
                tank_ready_time_str = _format_timedelta(duration_delta) 
            
            # Get the mix string for this specific tank_id AND cycle_num
            mix_string_for_cycle = tank_mix_map.get((tank_id, cycle_num), 'N/A') 
            
            # --- START FIX: Corrected the row_data list ---
            row_data = [
                tank_name_map.get(tank_id, f"Tank {tank_id}"),
                mix_string_for_cycle,
                fill_start.strftime('%d/%m/%y') if fill_start else '',
                fill_start.strftime('%H:%M') if fill_start else '',
                fill_end.strftime('%d/%m/%y') if fill_end else '',
                fill_end.strftime('%H:%M') if fill_end else '',
                settle_start.strftime('%d/%m/%y') if settle_start else '',
                settle_start.strftime('%H:%M') if settle_start else '',
                lab_start.strftime('%d/%m/%y') if lab_start else '',
                lab_start.strftime('%H:%M') if lab_start else '',
                ready_time.strftime('%d/%m/%y') if ready_time else '',
                ready_time.strftime('%H:%M') if ready_time else '',
                tank_ready_time_str  # <--- This is the correct final value
            ]
            # --- END FIX ---
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
            current_row += 1
        # --- End Iteration ---

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 5, 60)

        # Add borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None: cell.border = thin_border
        return True

    except Exception as e:
        print(f"Error creating sequence summary: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
def _create_tank_cargo_filling_sheet(wb, results):
    """Create Sheet: Tank x Cargo Filling Timeline"""
    try:
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        import re
        from datetime import datetime

        ws = wb.create_sheet("tank_cargo_filling")
        
        simulation_log = results.get('simulation_log', [])
        if not simulation_log:
            return False

        tanks = set()
        cargos = set()
        
        cargo_first_seen = {}
        tank_first_seen = {}
        
        fill_starts = {} 
        fill_ends = {}   
        fill_final_ends = {} 

        for log_entry in simulation_log:
            event = log_entry.get("Event", "")
            cargo_name = log_entry.get("Cargo", "")
            timestamp_str = log_entry.get("Timestamp", "")
            tank_str = log_entry.get("Tank", "")

            if not cargo_name or not tank_str:
                continue
            
            # Handle cycle numbers in event names - strip only the cycle number suffix
            event_base = event.rsplit('_', 1)[0] if '_' in event and event.split('_')[-1].isdigit() else event
            
            if event_base not in ["FILL_START", "FILL_START_FIRST", "FILL_END", "FILL_FINAL_END"]:
                continue
            
            try:
                timestamp_dt = datetime.strptime(timestamp_str, "%d/%m/%Y %H:%M")
                match = re.search(r'\d+', tank_str)
                if not match:
                    continue
                tank_id = int(match.group(0))
            except (ValueError, TypeError):
                continue

            key = (tank_id, cargo_name)
            
            tanks.add(tank_id)
            cargos.add(cargo_name)
            
            if cargo_name not in cargo_first_seen:
                cargo_first_seen[cargo_name] = timestamp_dt

            if event_base in ("FILL_START", "FILL_START_FIRST"):
                if tank_id not in tank_first_seen:
                    tank_first_seen[tank_id] = timestamp_dt
                
                if key not in fill_starts or timestamp_dt < fill_starts[key]:
                    fill_starts[key] = timestamp_dt
            
            if event_base == "FILL_END":
                if key not in fill_ends or timestamp_dt > fill_ends[key]:
                    fill_ends[key] = timestamp_dt
            
            if event_base == "FILL_FINAL_END":
                if key not in fill_final_ends or timestamp_dt > fill_final_ends[key]:
                    fill_final_ends[key] = timestamp_dt
        
        if not tank_first_seen:
            ws.cell(row=1, column=1, value="No fill data found.")
            return True

        sorted_tanks = sorted(list(tanks), key=lambda t: tank_first_seen.get(t, datetime.max)) 
        sorted_cargos = sorted(list(cargos), key=lambda c: cargo_first_seen.get(c, datetime.max))
        num_cargos = len(sorted_cargos)
        
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        cell = ws.cell(row=1, column=1, value="TANK")
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        ws.row_dimensions[1].height = 20
        
        current_col = 2
        for idx, cargo_name in enumerate(sorted_cargos):
            num_sub_cols = 2
            col_start = current_col
            col_end = current_col + num_sub_cols - 1
            
            cell = ws.cell(row=1, column=col_start, value=cargo_name)
            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            
            for c in range(col_start + 1, col_end + 1):
                ws.cell(row=1, column=c).fill = header_fill
            current_col = col_end + 1
            
        max_col_idx = current_col - 1

        ws.row_dimensions[2].height = 20
        cell = ws.cell(row=2, column=1, value="") 
        cell.fill = header_fill
        
        current_col = 2
        for idx, cargo_name in enumerate(sorted_cargos):
            is_first_cargo = (idx == 0)
            is_last_cargo = (idx == num_cargos - 1)
            
            if is_first_cargo:
                headers = ["FILL_START_FIRST", "FILL_END"]
            elif is_last_cargo:
                headers = ["FILL_START", "FILL_FINAL_END"]
            else:
                headers = ["FILL_START", "FILL_END"]
            
            for header in headers:
                cell = ws.cell(row=2, column=current_col, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.fill = header_fill
                current_col += 1

        current_row = 3
        for tank_id in sorted_tanks: 
            cell = ws.cell(row=current_row, column=1, value=f"Tank {tank_id}") 
            cell.font = bold_font
            cell.alignment = center_align
            
            current_col = 2
            for idx, cargo_name in enumerate(sorted_cargos):
                is_last_cargo = (idx == num_cargos - 1)
                key = (tank_id, cargo_name)

                start_dt = fill_starts.get(key)
                start_str = start_dt.strftime("%d/%m %H:%M") if start_dt else ""
                ws.cell(row=current_row, column=current_col, value=start_str).alignment = center_align
                
                end_dt = None
                if is_last_cargo:
                    end_dt = fill_final_ends.get(key)
                    if not end_dt:
                        end_dt = fill_ends.get(key)
                else:
                    end_dt = fill_ends.get(key)
                    final_end_dt = fill_final_ends.get(key)
                    if final_end_dt and (not end_dt or final_end_dt > end_dt):
                        end_dt = final_end_dt
                        
                end_str = end_dt.strftime("%d/%m %H:%M") if end_dt else ""
                ws.cell(row=current_row, column=current_col + 1, value=end_str).alignment = center_align
                current_col += 2

            current_row += 1

        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, max_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col_idx):
            for cell in row:
                cell.border = thin_border
        
        ws.freeze_panes = 'B3'
        
        return True

    except Exception as e:
        print(f"Error creating tank_cargo_filling sheet: {e}")
        import traceback
        traceback.print_exc()
        return False
def _create_tank_filling_volumes_sheet(wb, results):
    """Create the Tank Filling Volumes sheet"""
    try:
        ws = wb.create_sheet("Tank Filling Volumes")

        daily_discharge_log = results.get('daily_discharge_log', [])
        
        # DEBUG
        
        current_row = 1
        ws.cell(row=current_row, column=1, value="DAILY CARGO DISCHARGE").font = Font(bold=True, size=14)
        current_row += 2
        
        headers = ['DATE', 'CARGO', 'CRUDE_TYPE', 'DISCHARGE (bbls)', 'TANK', 'VOL_FILLED (bbls)']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=current_row, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            c.alignment = Alignment(horizontal='center')
        current_row += 1
    
        if not daily_discharge_log:
            # No data available
            ws.cell(row=current_row, column=1, value="No discharge data available")
            return True
    
        consolidated_data = {}
        for entry in daily_discharge_log:
            key = (
                entry.get('date', 'Unknown'), 
                entry.get('cargo_type', 'Unknown'),
                entry.get('crude_type', 'N/A'), 
                entry.get('tank_id', 0)
            )
            if key not in consolidated_data:
                consolidated_data[key] = 0
            consolidated_data[key] += entry.get('volume_filled', 0)
            
        report_events = []
        for (date, cargo_type, crude_type, tank_id), volume in consolidated_data.items():
            report_events.append({
                'date': date, 'cargo_type': cargo_type,
                'crude_type': crude_type, 'tank_id': tank_id,
                'volume_filled': volume
            })
            
        def get_sort_datetime(event):
            try:
                return datetime.strptime(event['date'], '%d/%m/%y')
            except (ValueError, TypeError):
                return datetime.min
    
        report_events.sort(key=get_sort_datetime)
    
        if not report_events:
            ws.cell(row=current_row, column=1, value="No discharge events to display")
            return True

        events_iterator = iter(report_events)
        current_event = next(events_iterator, None)
        operation_subtotal = 0
    
        while current_event:
            tank_id = current_event['tank_id']
            tank_display_name = f"Tank {tank_id}"
        
            operation_subtotal += current_event['volume_filled']
        
            ws.cell(row=current_row, column=1, value=current_event['date']).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=2, value=current_event['cargo_type']).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=3, value=current_event.get('crude_type', 'N/A')).alignment = Alignment(horizontal='center')
        
            discharge_cell = ws.cell(row=current_row, column=4, value=current_event['volume_filled'])
            discharge_cell.number_format = '#,##0'
            discharge_cell.alignment = Alignment(horizontal='center')

            ws.cell(row=current_row, column=5, value=tank_display_name).alignment = Alignment(horizontal='center')

            vol_filled_cell = ws.cell(row=current_row, column=6, value=operation_subtotal)
            vol_filled_cell.number_format = '#,##0'
            vol_filled_cell.alignment = Alignment(horizontal='center')
        
            current_row += 1
            next_event = next(events_iterator, None)
        
            operation_ended = False
            if next_event is None or next_event['tank_id'] != tank_id or (get_sort_datetime(next_event) - get_sort_datetime(current_event)).days > 1:
                operation_ended = True

            if operation_ended:
                subtotal_cell = ws.cell(row=current_row, column=5, value=f"Subtotal {tank_display_name}")
                subtotal_cell.font = Font(bold=True)
                subtotal_cell.alignment = Alignment(horizontal='right')
            
                volume_cell = ws.cell(row=current_row, column=6, value=operation_subtotal)
                volume_cell.font = Font(bold=True)
                volume_cell.number_format = '#,##0'
                volume_cell.alignment = Alignment(horizontal='center')
                current_row += 1
                operation_subtotal = 0
        
            current_event = next_event

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, 12), 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=current_row-1):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
        return True
    
    except Exception as e:
        print(f"Error creating tank filling volumes sheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def register_routes(app):
    """Register all routes with the Flask app"""
    
    # Set secret key if not already set
    if not app.secret_key:
        app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'crude-scheduler-secret-key-change-in-production')

    @app.route('/')
    def root():
        """Always start at the login page."""
        return redirect(url_for('login_page'))

    @app.route('/login', methods=['GET', 'POST'])
    def login_page():
        # Handles login logic
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            if username == APP_USERNAME and password == APP_PASSWORD:
                return redirect(url_for('index'))
            else:
                return render_template('login.html', message='Invalid username or password')
        return render_template('login.html')

    @app.route('/index')
    def index():
        """Displays the main scheduler page."""
        return render_template('index.html')

    @app.route('/api/simulate', methods=['POST'])
    def simulate():
        """
        API endpoint that runs the simulation.
        If optimization is requested, it runs the solver first and passes the
        results to the main simulator.
        """
        try:
            params = request.json
            if not params:
                return jsonify({'error': 'No data provided'}), 400

            # Check if an optimized schedule should be used.
            if params.get('use_optimized_schedule'):

                # 1. Call the solver to generate the schedule.
                solver_results = optimize_crude_mix_schedule(params)

                # 2. Check for solver errors and return a clear message if it fails.
                if not solver_results or not solver_results.get('success'):
                    error_message = solver_results.get('error', 'Solver failed to produce a valid plan.')
                    return jsonify({'success': False, 'error': error_message}), 400

                # 3. Add the solver's output to the parameters for the main simulator.
                params['solver_results'] = solver_results
            
            # --- PREPARE SIMULATOR CONFIG (CFG) ---
            # This block is the same as your old routes1.py, it prepares the data.
            try:
                # Calculate Usable Volume
                unusable_part = safe_float(params.get('defaultDeadBottom', 10000), 10000, 'defaultDeadBottom') + \
                               (safe_float(params.get('bufferVolume', 500), 500, 'bufferVolume') / 2.0)
                usable_volume = safe_float(params.get('tankCapacity', 600000), 600000, 'tankCapacity') - unusable_part

                initial_tank_levels = {}
                num_tanks = safe_int(params.get('numTanks'), param_name='numTanks')

                # Add validation for number of tanks
                if num_tanks <= 0:
                    return jsonify({'success': False, 'error': 'Number of tanks must be greater than zero.'}), 400

                for i in range(1, num_tanks + 1):
                    tank_level = safe_float(params.get(f'tank{i}Level', usable_volume), usable_volume, f'tank{i}Level')
                    initial_tank_levels[i] = tank_level

                # Prepare the configuration dictionary for the Simulator
                cfg = {
                    "processing_rate": safe_float(params.get('processingRate'), param_name='processingRate'),
                    "num_tanks": num_tanks,
                    "initial_tank_levels": initial_tank_levels,
                    "start_dt": datetime.strptime(params['crudeProcessingDate'], "%Y-%m-%dT%H:%M"),
                    "usable_per_tank": usable_volume,
                    "settling_days": safe_float(params.get('settlingTime'), param_name='settlingTime'),
                    "lab_hours": safe_float(params.get('labTestingDays', 0), param_name='labTestingDays') * 24.0,
                    "discharge_rate": safe_float(params.get('pumpingRate'), param_name='pumpingRate'),
                    "dead_bottom": safe_float(params.get('defaultDeadBottom', 10000), 10000, 'defaultDeadBottom'),
                    "buffer_volume": safe_float(params.get('bufferVolume', 500), 500, 'bufferVolume'),
                    "min_ready_tanks": safe_int(params.get('minReadyTanks', 2), 2, 'minReadyTanks'),
                    "first_cargo_min_ready": safe_int(params.get('firstCargoMinReady', 8), 8, 'firstCargoMinReady'),
                    "first_cargo_max_ready": safe_int(params.get('firstCargoMaxReady', 9), 9, 'firstCargoMaxReady'),
                    "tank_gap_hours": safe_float(params.get('tankGapHours', 0.0), 0.0, 'tankGapHours'),
                    "berth_gap_hours_min": safe_float(params.get('berth_gap_hours_min', 0), 0, 'berth_gap_hours_min'),
                    "berth_gap_hours_max": safe_float(params.get('berth_gap_hours_max', 0), 0, 'berth_gap_hours_max'),
                    "preDischargeDays": safe_float(params.get('preDischargeDays', 0), param_name='preDischargeDays'),
                    "tankFillGapHours": safe_float(params.get('tankFillGapHours', 0.0), 0.0, 'tankFillGapHours'),
                    "horizon_days": safe_float(params.get('schedulingWindow'), param_name='schedulingWindow'),
                    "snapshot_interval_minutes": safe_int(params.get('snapshotIntervalMinutes', 30), 30, 'snapshotIntervalMinutes'),
                    "cargo_defs": {
                        "VLCC": safe_float(params.get('vlccCapacity', 0), 0, 'vlccCapacity'),
                        "SUEZ": safe_float(params.get('suezmaxCapacity', 0), 0, 'suezmaxCapacity'),
                        "AFRA": safe_float(params.get('aframaxCapacity', 0), 0, 'aframaxCapacity'),
                        "PANA": safe_float(params.get('panamaxCapacity', 0), 0, 'panamaxCapacity'),
                        "HANDY": safe_float(params.get('handymaxCapacity', 0), 0, 'handymaxCapacity'),
                    },
                    "use_optimized_schedule": params.get('use_optimized_schedule', False),
                    "solver_results": params.get('solver_results', None)
                }

            except (KeyError, ValueError) as e:
                return jsonify({'error': f'Invalid or missing parameter: {str(e)}'}), 400

            # --- RUN THE SIMULATION ---
            sim = Simulator(cfg)
            sim.run()

            if sim.infeasible:
                return jsonify({'success': False, 'error': 'Simulation Infeasible', 'message': sim.infeasible_reason}), 400

            sim.generate_cargo_report()
            sim.daily_log_rows.sort(key=lambda x: datetime.strptime(x["Timestamp"], "%d/%m/%Y %H:%M"))
            
            # NOTE: We skip sim.save_csvs() because the function is complex and handled below.
            # We ONLY need to generate the cargo report and sort the log before saving.

            # --- START FINAL STREAMLIT API FILE SAVES (CLEANED BLOCK) ---
            alerts = []
            for log_entry in sim.daily_log_rows:
                if log_entry.get("Level", "").lower() in ["danger", "warning"]:
                    day_num = (datetime.strptime(log_entry['Timestamp'], "%d/%m/%Y %H:%M") - cfg['start_dt']).days + 1
                    alerts.append({"day": day_num, "type": log_entry.get("Level").lower(), "message": log_entry.get("Message")})

            # --- START FINAL CRITICAL FIX BLOCK ---
            try:
                if sim.daily_summary_rows:
                    # CORRECT: Convert list of dicts to DataFrame
                    pd.DataFrame(sim.daily_summary_rows).to_csv("/tmp/daily_summary.csv", index=False)
                
                if sim.daily_log_rows:
                    # CORRECT: Convert list of dicts to DataFrame
                    pd.DataFrame(sim.daily_log_rows).to_csv("/tmp/simulation_log.csv", index=False)
                    
                if sim.cargo_report_rows:
                    # CORRECT: Convert list of dicts to DataFrame
                    pd.DataFrame(sim.cargo_report_rows).to_csv("/tmp/cargo_report.csv", index=False)
                
                if sim.snapshot_log:
                    # CORRECT: Convert list of dicts to DataFrame
                    pd.DataFrame(sim.snapshot_log).to_csv("/tmp/tank_snapshots.csv", index=False)

            except Exception as e:
                print(f"Error saving all Streamlit API files: {e}")
                return jsonify({'error': f'Failed to save necessary CSV files: {str(e)}'}), 500
            
            # --- END FINAL CRITICAL FIX BLOCK ---

            # Create the simple map Streamlit needs for downloads
            csv_download_urls = {
                "daily_summary.csv": "/api/get_results",
                "simulation_log.csv": "/download/simulation_log.csv",
                "cargo_report.csv": "/download/cargo_report.csv",
                "tank_snapshots.csv": "/download/tank_snapshots.csv"
            }

            # NOTE: The manual download logic is handled in the frontend via main.js
            
            return jsonify({
                'success': True,
                'csv_files': csv_download_urls,
                'daily_summary': sim.daily_summary_rows,
                'cargo_report': sim.cargo_report_rows,
                'simulation_log': sim.daily_log_rows,
                'alerts': alerts,
                'simulation_data': sim.daily_summary_rows,
                'feeding_events_log': sim.feeding_events_log, 
                'filling_events_log': sim.filling_events_log, 
                'daily_discharge_log': sim.daily_discharge_log 
            })

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"\nERROR: Exception in simulate endpoint:\n{error_details}")
            return jsonify({'error': 'An unexpected server error occurred.', 'details': str(e)}), 500

 
    @app.route('/api/export_tank_status', methods=['POST'])
    def export_tank_status():
        """Export sequence report with Sequence Summary, Tank Filling Volumes, and Tank-Cargo-Filling sheets"""
        try:
            # CREATE DEEP COPY
            results = copy.deepcopy(request.json)
        
            # Create a new workbook
            wb = Workbook()
        
            # Remove the default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
            # Create the sequence summary sheet
            sequence_success = _create_sequence_summary_sheets(wb, results)
        
            # Create the tank filling volumes sheet
            volume_success = _create_tank_filling_volumes_sheet(wb, results)
        
            # --- ADD THIS LINE ---
            # Create the tank cargo filling sheet
            tank_cargo_success = _create_tank_cargo_filling_sheet(wb, results)
            # --- END ADD ---
        
            if not sequence_success:
                return jsonify({'error': 'Failed to create sequence summary'}), 400
            
            if not volume_success:
                return jsonify({'error': 'Failed to create tank filling volumes sheet'}), 400
            
            # --- ADD THIS CHECK ---
            if not tank_cargo_success:
                return jsonify({'error': 'Failed to create tank cargo filling sheet'}), 400
            # --- END ADD ---
        
            # Generate download filename with timestamp
            timestamp_str = datetime.now().strftime('%d-%b-%Y_%H-%M-%S')
            download_filename = f"sequence_report_{timestamp_str}.xlsx"
        
            # Create temporary file with context manager for auto-cleanup
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
        
            def remove_file():
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
                
            # Send file and schedule cleanup
            response = send_file(
                tmp_file_path,
                as_attachment=True,
                download_name=download_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
            )
    
            # Clean up temp file immediately
            remove_file()
        
            return response
        
        except Exception as e:
            print(f"Error in export_tank_status: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/export_charts', methods=['POST'])
    def export_charts():
        """Export comprehensive charts workbook with 3 sheets"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            
            # Validate request data
            if not request.json:
                print("ERROR: No JSON data received")
                return jsonify({'error': 'No data provided'}), 400
            
            # Validate required keys
            required_keys = ['simulation_data', 'cargo_report', 'simulation_log']
            missing_keys = [key for key in required_keys if key not in request.json]
            if missing_keys:
                print(f"ERROR: Missing required keys: {missing_keys}")
                return jsonify({'error': f'Missing required data: {", ".join(missing_keys)}'}), 400
            
            # Validate data is not empty
            if not request.json.get('simulation_data'):
                print("ERROR: simulation_data is empty")
                return jsonify({'error': 'simulation_data is empty'}), 400
            
            if not request.json.get('cargo_report'):
                print("ERROR: cargo_report is empty")
                return jsonify({'error': 'cargo_report is empty'}), 400
            
            if not request.json.get('simulation_log'):
                print("ERROR: simulation_log is empty")
                return jsonify({'error': 'simulation_log is empty'}), 400
            
            # CREATE DEEP COPY to protect original data
            results = copy.deepcopy(request.json)
            
            
            # Create workbook with timestamp in workbook name
            wb = Workbook()
            timestamp_str = datetime.now().strftime('%d-%b-%Y %H:%M:%S')
            wb.properties.title = f"charts {timestamp_str}"
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create all 3 sheets
            success_results = {}
            
            success_results['sheet1'] = _create_simulation_log_sheet(wb, results)
            success_results['sheet2'] = _create_daily_summary_sheet(wb, results)
            success_results['sheet3'] = _create_cargo_arrivals_sheet(wb, results)
            success_results['sheet3_chart'] = _create_certified_stock_chart_sheet(wb, results)
            
            # Check if any sheet creation failed
            failed_sheets = [k for k, v in success_results.items() if not v]
            if failed_sheets:
                return jsonify({'error': f'Failed to create sheets: {", ".join(failed_sheets)}'}), 400
            
            # Create a proper temporary file that gets deleted immediately after sending
            
            # Generate download filename with timestamp
            timestamp_str_file = datetime.now().strftime('%d-%b-%Y_%H-%M-%S')
            download_filename = f"charts_report_{timestamp_str_file}.xlsx"
            
            # Create temporary file with context manager for auto-cleanup
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
            
            def remove_file():
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
            
            # Send file and schedule cleanup
            response = send_file(
                tmp_file_path,
                as_attachment=True,
                download_name=download_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Clean up temp file immediately
            remove_file()
            
            return response
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"ERROR: Exception in export_charts: {str(e)}")
            print(f"ERROR: Full traceback:\n{error_details}")
            return jsonify({'error': f'Export failed: {str(e)}', 'details': error_details}), 500

    @app.route('/api/save_inputs', methods=['POST'])
    def save_inputs():
        """Save user inputs (dummy endpoint for now)"""
        return jsonify({'success': True, 'message': 'Inputs saved'})

    @app.route('/api/load_inputs', methods=['GET'])
    def load_inputs():
        """Load user inputs (dummy endpoint for now)"""
        return jsonify({})

    @app.route('/api/optimize_crude_mix', methods=['POST'])
    def optimize_crude_mix():
        """API endpoint to optimize crude mix schedule using the solver."""
        try:
            params = request.json
            
            if not params:
                return jsonify({'error': 'No data provided'}), 400


            # --- Early Validation ---
            crude_names = params.get('crude_names', [])
            crude_percentages = params.get('crude_percentages', [])

            if not crude_names or not crude_percentages:
                return jsonify({
                    'success': False, 'error': 'Missing crude mix data',
                    'details': 'crude_names and crude_percentages are required.'
                }), 400
            
            if len(crude_names) != len(crude_percentages):
                return jsonify({
                    'success': False, 'error': 'Crude mix data mismatch',
                    'details': 'The number of crude names does not match the number of percentages.'
                }), 400

            # --- Call the Actual Solver ---
            # The solver function from solver.py will handle the main logic and printing.
            solver_results = optimize_crude_mix_schedule(params)


            if solver_results.get('success'):
                return jsonify({
                    'success': True,
                    'optimization_results': solver_results
                })
            else:
                # If the solver failed, return its error message directly.
                return jsonify(solver_results)

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"\nERROR: Exception in optimize_crude_mix endpoint:")
            print(error_details)
            return jsonify({
                'success': False,
                'error': 'An unexpected error occurred during optimization.',
                'details': str(e),
                'traceback': error_details
            }), 500

    @app.route('/api/export_solver_report', methods=['POST'])
    def export_solver_report():
        """Export the detailed console output from the solver as a text file."""
        try:
            solver_results = request.json
            
            if not solver_results:
                return jsonify({'error': 'No optimization results provided'}), 400
            
            # The solver.py file conveniently provides a formatted console output log.
            console_output = solver_results.get('console_output', [])
            if not console_output:
                report_content = "No console output was generated by the solver."
            else:
                report_content = "\n".join(console_output)

            # Create an in-memory text file
            output = io.BytesIO()
            output.write(report_content.encode('utf-8'))
            output.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"solver_report_{timestamp}.txt"
            
            return send_file(
                output,
                mimetype='text/plain',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Failed to export solver report: {str(e)}'}), 500
    
    # routes.py (NEW /download/<filename> endpoint using send_from_directory)

    @app.route('/api/get_results', methods=['GET'])
    def get_simulation_results():
        """
        NEW endpoint for Streamlit. 
        It sends the file from the fixed path and does NOT delete it.
        """
        if not os.path.exists(SUMMARY_FILE_PATH):
            return jsonify({"error": "File not found. Please run the simulation first."}), 404
        
        # This sends the file from the Flask app's disk
        return send_file(
            SUMMARY_FILE_PATH,
            mimetype='text/csv',
            download_name='daily_summary.csv',
            as_attachment=True
        )

    @app.route("/download/<filename>")
    def download_file(filename):
        """
        Download simulation results. Never deletes Streamlit dashboard files.
        """
        directory_path = "/tmp"
        
        # --- CRITICAL: Files that Streamlit needs must NEVER be deleted ---
        STREAMLIT_FILES_TO_KEEP = [
            "simulation_log.csv", 
            "cargo_report.csv", 
            "tank_snapshots.csv",  
            "daily_summary.csv"    
        ]
        
        try:
            # Security check
            if ".." in filename or "/" in filename:
                print(f"[SECURITY] Invalid filename requested: {filename}")
                return jsonify({"error": "Invalid filename requested"}), 400

            file_path = os.path.join(directory_path, filename)
            
            if os.path.exists(file_path):
                # Check file size and modification time
                file_size = os.path.getsize(file_path)
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                

            if not os.path.exists(file_path):
                print(f"[ERROR] File not found: {file_path}")
                return jsonify({"error": "File not found on server"}), 404
            
            # NEVER delete Streamlit files
            should_delete = filename not in STREAMLIT_FILES_TO_KEEP
            
            response = send_from_directory(
                directory_path, 
                filename, 
                as_attachment=True, 
                download_name=filename
            )
            
            # Only schedule cleanup for non-Streamlit files
            if should_delete:
                print(f"[INFO] Will delete {filename} after sending")
                @response.call_on_close
                def cleanup_file():
                    try:
                        os.remove(file_path)
                        print(f"[CLEANUP] Deleted {file_path}")
                    except Exception as e:
                        print(f"[CLEANUP ERROR] Failed to delete {file_path}: {e}")
            else:
                print(f"[INFO] Keeping {filename} (Streamlit dashboard file)")

            return response
            
        except Exception as e:
            print(f"[ERROR] Exception in download_file: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": "An unexpected error occurred during file transfer."}), 500