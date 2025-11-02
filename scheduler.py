#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import math
import csv
import os
import re
import random

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Excel auto-formatting will not be available.")
    print("Install with: pip install openpyxl")

# Import solver plan manager for optimized scheduling
try:
    from scheduler_solver import SolverPlanManager
    SOLVER_AVAILABLE = True
except ImportError:
    SOLVER_AVAILABLE = False
    print("WARNING: scheduler_solver.py not found. Solver-based scheduling unavailable.")

# ------------------------- CONSTANTS / STATES -------------------------
READY, FEEDING, EMPTY, FILLING, FILLED, SETTLING, LAB, SUSPENDED = (
    "READY", "FEEDING", "EMPTY", "FILLING", "FILLED", "SETTLING", "LAB", "SUSPENDED"
)

# ------------------------- SIMULATOR -------------------------
class Simulator:
    def __init__(self, cfg):
        # Store original config for solver access
        self.cfg = cfg
        
        # Inputs
        self.rate_day = cfg["processing_rate"]  # FIXED processing rate
        self.rate_hour = self.rate_day / 24.0  # FIXED hourly rate
        self.N = cfg["num_tanks"]
        self.start = cfg["start_dt"]
        self.usable = cfg["usable_per_tank"]
        self.dead_bottom = cfg["dead_bottom"]
        self.buffer_volume = cfg["buffer_volume"]
        self.unusable_per_tank = self.dead_bottom + (self.buffer_volume / 2.0)
        self.feeding_events_log = []
        self.filling_events_log = []
        self.daily_discharge_log = []
        self.snapshot_log = []

        self.settle_hours = cfg["settling_days"] * 24.0
        self.lab_hours = cfg["lab_hours"]
        self.discharge_rate = cfg["discharge_rate"]
        self.fill_hours = self.usable / max(self.discharge_rate, 1e-6)
        self.enabled_cargos: Dict[str, int] = cfg["cargo_defs"]
        self.horizon_days = cfg["horizon_days"]
        self.min_ready_tanks = cfg["min_ready_tanks"]
        self.first_cargo_min_ready = cfg.get("first_cargo_min_ready", 8)
        self.first_cargo_max_ready = cfg.get("first_cargo_max_ready", 9)
        
        # Snapshot interval in minutes (from config or default to 30)
        self.snapshot_interval_minutes = cfg.get("snapshot_interval_minutes", 30)
        
        self.berth_gap_hours_min = float(cfg.get("berth_gap_hours_min", 0.0))
        self.berth_gap_hours_max = float(cfg.get("berth_gap_hours_max", 0.0))
        self.fill_delay_hours = float(cfg.get("preDischargeDays", 0)) * 24.0
        self.tank_gap_hours = float(cfg.get("tank_gap_hours", 0.0))
        self.ready_for_fill_at: Dict[int, Optional[datetime]] = {i: datetime.min for i in range(1, self.N + 1)}
        self.tank_fill_gap_hours = float(cfg.get("tankFillGapHours", 0.0))

        # Tank state - all tanks start with exactly usable volume
        self.state: Dict[int, str] = {i: READY for i in range(1, self.N + 1)}
        
        # Track state changes with timestamps for accurate logging
        self.state_history: List[Tuple[datetime, int, str]] = []  # [(timestamp, tank_id, state)]

        initial_levels = cfg.get("initial_tank_levels", {})
        self.bbl: Dict[int, float] = {i: initial_levels.get(i, self.usable) for i in range(1, self.N + 1)}
        # Track daily consumption per tank
        self.daily_consumption: Dict[int, float] = {i: 0.0 for i in range(1, self.N + 1)}

        # FIX: Normalize initial levels to USABLE-ONLY and mark EMPTY accordingly
        for i in range(1, self.N + 1):
            gross = float(self.bbl[i] or 0.0)                   # some inputs include heel
            usable = max(0.0, gross - self.unusable_per_tank)   # store usable only
            self.bbl[i] = usable
            if usable <= 0.0:
                self.state[i] = EMPTY

        # FIX 2: Track tanks that started empty (should be filled first)
        self.initially_empty_tanks = [i for i in range(1, self.N + 1) if self.state[i] == EMPTY]
        # --- CORRECT INITIALIZATION FOR CYCLE COUNTER ---
        # Initialize cycle counter based on initial state
        self.tank_cycle_counter: Dict[int, int] = {i: 1 for i in range(1, self.N + 1)}
        # --- END FINAL CORRECT INITIALIZATION ---

        self.ready_at: Dict[int, Optional[datetime]] = {i: None for i in range(1, self.N + 1)}
                
        self.settle_end_at: Dict[int, Optional[datetime]] = {i: None for i in range(1, self.N + 1)}
        self.lab_start_at: Dict[int, Optional[datetime]] = {i: None for i in range(1, self.N + 1)}
        
        # Track initial volume when tank starts feeding (for accurate total draw calculation)
        self.feed_start_volume: Dict[int, float] = {i: 0.0 for i in range(1, self.N + 1)}
        self.tank_feed_start_time: Dict[int, Optional[datetime]] = {i: None for i in range(1, self.N + 1)}

        # --- crude mix tracking (used by scheduler_solver) ---
        self.tank_mix = {i: {} for i in range(1, self.N + 1)}       # crude bbl per tank
        self.tank_mix_pct = {i: {} for i in range(1, self.N + 1)}   # crude % per tank
        # --- end crude mix tracking ---

        # Track alert flags
        self.no_feed_alert_logged = False
        self.first_cargo_scheduled = False  # Track if first cargo has been scheduled
        self.tank_filled_first = set() 
        self.cargo_has_started_filling = set() # <-- ADD THIS NEW LINE

        # Start with Tank 1 feeding
        self.active = 1
        self._change_state(self.active, FEEDING, self.start)
        self.bbl[self.active] = min(self.bbl[self.active], self.usable)
        self.feed_start_volume[self.active] = self.bbl[self.active]  # Track starting volume
        
        
        # Berth management: {berth_id: {"free_at": datetime, "current_cargo": dict or None}}
        self.berths = {
            1: {"free_at": self.start, "current_cargo": None},
            2: {"free_at": self.start, "current_cargo": None}
        }

        # Filling control: only one tank can be filled at a time per cargo
        # {cargo_vessel_name: (tank_id, end_time, volume_to_fill)}
        self.active_fills: Dict[str, Tuple[int, datetime, float]] = {}

        # Cargo tracking with unique vessel names
        self.cargo_counter = {"VLCC": 0, "SUEZ": 0, "AFRA": 0, "PANA": 0, "HANDY": 0}
        self.cargos: List[Dict] = []  # Each cargo has unique vessel name

        # Track remaining volume per cargo
        self.cargo_remaining_volume: Dict[str, float] = {}

        # Outputs
        self.daily_log_rows: List[Dict] = []
        self.daily_summary_rows: List[Dict] = []
        self.cargo_report_rows: List[Dict] = []
        self.inventory_data: List[Tuple[datetime, float]] = []  # for chart
        
        self.infeasible = False
        self.infeasible_reason = ""
        
        # Tanks needed per cargo type
        self.tanks_needed_by_type: Dict[str, int] = {
            name: math.ceil(vol / self.usable) for name, vol in self.enabled_cargos.items()
        }

        # Initial log
        self._log_event(self.start, "Info", "SIM_START", None, None,
                        f"Simulation started with processing rate: {int(self.rate_day):,} bbl/day")
        self._log_event(self.start, "Info", "FEED_START", self.active, None,
                        f"Initial feeding starts from Tank {self.active}")
        self.tank_feed_start_time[self.active] = self.start
       
        self._log_event(self.start, "Info", "CONFIG", None, None,
                f"CONFIG: usable_per_tank={self.usable}, dead_bottom={self.dead_bottom}, buffer_volume={self.buffer_volume}, unusable={self.unusable_per_tank}")

        # ==================== SOLVER INITIALIZATION ====================
        # Initialize solver plan manager if available
        self.use_solver_plan = cfg.get('use_optimized_schedule', False)
        self.solver_plan_manager = None
        self.solver_results = cfg.get('solver_results', None)

        if self.use_solver_plan and SOLVER_AVAILABLE:
            
            self.solver_plan_manager = SolverPlanManager(self)
            
            # Pass the full config with solver results
            solver_init_params = cfg.copy()
            solver_initialized = self.solver_plan_manager.initialize_solver_plan(solver_init_params)
            
            if solver_initialized:
                self._log_event(self.start, "Info", "SOLVER_INIT", None, None,
                               "Solver-based optimization plan loaded successfully")
            else:
                self.use_solver_plan = False
                self._log_event(self.start, "Warning", "SOLVER_INIT_FAIL", None, None,
                               "Failed to load solver plan - using standard scheduling")
        else:
            if self.use_solver_plan and not SOLVER_AVAILABLE:
                self._log_event(self.start, "Warning", "SOLVER_UNAVAILABLE", None, None,
                               "Solver requested but module not found - using standard scheduling")
                self.use_solver_plan = False
        
        # Load solver cargos if solver is being used
        if self.use_solver_plan:
            self._load_solver_cargos()
        # ==================== END SOLVER INITIALIZATION ====================
    # scheduler.py (Fixed _load_solver_cargos method)

    def _load_solver_cargos(self):
        """Pre-load all cargos from solver plan during initialization"""
        if not self.use_solver_plan or not self.solver_plan_manager:
            return
        
        solver_results = self.solver_results or {}
        cargo_schedule = solver_results.get('cargo_schedule', [])
        
        if not cargo_schedule:
            return
        
        
        # --- FIX: Remove last_arrival_time tracking and gap calculation for pre-loading ---
        # All solver cargos are considered ready to arrive at self.start.
        # last_arrival_time: Dict[int, datetime] = { 1: self.start, 2: self.start }
        # --- END FIX ---
        
        berth_id = 1  # Start with berth 1, alternate if needed
        
        # CRITICAL FIX: Load ALL cargos from solver
        for idx, cargo_data in enumerate(cargo_schedule):
            vessel_name = cargo_data.get('vessel_name', f"SOLVER-{cargo_data['cargo_id']}")
            
            # --- FIX: Set arrival time directly to self.start for all solver cargos ---
            arrival_time = self.start
            # --- END FIX ---
            
            cargo = {
                "vessel_name": vessel_name,
                "cargo_type": cargo_data.get('type', 'UNKNOWN'),
                "cargo_id": cargo_data['cargo_id'],
                "crude_type": cargo_data.get('crude_name', 'Unknown'),
                "berth": berth_id,
                "arrival": arrival_time, # Set to self.start
                "fill_start": arrival_time + timedelta(hours=self.fill_delay_hours),
                "volume": cargo_data['size'],
                "tanks_needed": math.ceil(cargo_data['size'] / self.usable),
                "tanks_started": 0,
                "tanks_done": 0,
                "discharge_start": None,
                "discharge_end": None,
                "tank_fills": [],
                "dispatched": False
            }
            
            self.cargos.append(cargo)
            self.cargo_remaining_volume[vessel_name] = cargo['volume']
            
#             print(f"  {cargo_data['cargo_id']:3d}. {vessel_name:20s} (Berth {berth_id}) - Arrival: {arrival_time.strftime('%d/%m %H:%M')}")
            
            # --- FIX: Remove update of last_arrival_time ---
            # last_arrival_time[berth_id] = arrival_time 
            
            # Alternate berths for the next cargo
            berth_id = 2 if berth_id == 1 else 1
        

    # ------------------------- LOGGING -------------------------
    def _change_state(self, tank_id: int, new_state: str, when: datetime):
        """Change tank state and record in history"""
        self.state[tank_id] = new_state
        self.state_history.append((when, tank_id, new_state))
    
    def _get_state_at_time(self, ts: datetime) -> Dict[int, str]:
        """Get tank states as they were at a specific timestamp"""
        # Start with initial states (all READY except EMPTY ones)
        states = {}
        for i in range(1, self.N + 1):
            if i in self.initially_empty_tanks:
                states[i] = EMPTY
            else:
                states[i] = READY
        
        # Apply state changes in chronological order up to this timestamp
        for change_time, tank_id, new_state in self.state_history:
            if change_time <= ts:
                states[tank_id] = new_state
            else:
                break  # History is chronological, so we can stop
        
        return states
    
    def _log_event(self, ts: datetime, level: str, event: str,
                   tank_id: Optional[int], cargo: Optional[str], message: str,
                   state_override: Optional[Dict[int, str]] = None):
        """Logs an event, appending cycle number to relevant event names."""

        event_name_to_log = event # Default to original event name

        # --- CORRECT LOGIC TO APPEND CYCLE NUMBER ---
        # List of events that track filling/processing cycles
        cycle_events = {"FILL_START_FIRST", "FILL_FINAL_END", "SETTLING_START", "SETTLING_END", "READY"}

        if event in cycle_events and tank_id is not None:
            # Get the current cycle number for this tank
            # Use default 1 if somehow missing, though it shouldn't be with correct init
            cycle_num = self.tank_cycle_counter.get(tank_id, 1)
            # Append the cycle number
            event_name_to_log = f"{event}_{cycle_num}"
        # --- END CORRECT LOGIC ---

        # Build tank status snapshot - use override if provided, else get state at this timestamp
        if state_override is not None:
            tank_status = {f"Tank{i}": state_override.get(i, self.state[i]) for i in range(1, self.N + 1)}
        else:
            states_at_ts = self._get_state_at_time(ts)
            tank_status = {f"Tank{i}": states_at_ts[i] for i in range(1, self.N + 1)}

        row = {
            "Timestamp": ts.strftime("%d/%m/%Y %H:%M"),
            "Level": level,
            "Event": event_name_to_log, # <<< Use the potentially modified event name
            "Tank": f"Tank {tank_id}" if tank_id else "",
            "Cargo": cargo or "",
            "Message": message
        }
        row.update(tank_status)
        self.daily_log_rows.append(row)

    # ------------------------- UTILITIES -------------------------
    def _count_state(self, target: str) -> int:
        return sum(1 for s in self.state.values() if s == target)

    def _sum_stock_ready_and_feeding(self) -> float:
        return sum(self.bbl[i] for i in range(1, self.N + 1) if self.state[i] in (READY, FEEDING))
    

    def _predict_next_tank_empty(self, now: datetime) -> Optional[timedelta]:
        """Predict when the next tank will become empty"""
        if self.active == 0 or self.state.get(self.active) != FEEDING:
            return None
        
        # Current feeding tank's time to empty
        available = self.bbl[self.active]
        hours_to_empty = available / self.rate_hour
        
        # Look ahead through READY tanks in sequence
        ready_tanks = [i for i in range(1, self.N + 1) if self.state[i] == READY]
        total_hours = hours_to_empty
        
        for _ in ready_tanks:
            total_hours += self.usable / self.rate_hour
        
        return timedelta(hours=total_hours)


    # ------------------------- CARGO SCHEDULING -------------------------
    def schedule_cargos(self, now: datetime):
        """Schedule cargos - uses solver plan if available, otherwise standard logic"""
        
        # If solver plan is active, delegate to solver plan manager
        if self.use_solver_plan and self.solver_plan_manager:
            return self._schedule_cargos_with_solver(now)
        else:
            return self._schedule_cargos_standard(now)

    def _schedule_cargos_with_solver(self, now: datetime):
        """Schedule cargos following solver's predetermined plan - cargo arrives AFTER berth is free"""
        
        for cargo in self.cargos:
            if cargo.get('dispatched', False) == False:
                berth = self.berths[cargo['berth']]
                
               
                # 1. Calculate a random gap for this specific cargo
                random_gap_hours = random.uniform(self.berth_gap_hours_min, self.berth_gap_hours_max)

                # 2. Calculate when this cargo *can* arrive.
                #    It's the time the berth was last free + the required gap.
                earliest_arrival_time = berth["free_at"] + timedelta(hours=random_gap_hours)
                # --- END MODIFICATION ---

                # 3. Check if the berth is free AND we are at (or past) the calculated arrival time
                if berth["current_cargo"] is None and now >= earliest_arrival_time:
                    
                    cargo['dispatched'] = True
                    berth["current_cargo"] = cargo
                    
                    # 3. Update the cargo's times to the *actual* calculated times
                    actual_arrival = earliest_arrival_time
                    cargo['arrival'] = actual_arrival
                    cargo['fill_start'] = actual_arrival + timedelta(hours=self.fill_delay_hours)
                    
                    # 4. Do NOT update berth["free_at"] here. This is only done
                    #    when this cargo *finishes* (in _maybe_finish_fill).

                    # --- END FIX ---

                    # Log arrival event ONLY ONCE
                    if not cargo.get('arrival_logged', False):
                        self._log_event(
                            actual_arrival,
                            "Success",
                            "ARRIVAL",
                            None,
                            cargo["vessel_name"],
                            f"BERTH {cargo['berth']} CARGO ARRIVED. Fill starts at {cargo['fill_start'].strftime('%d/%m %H:%M')}"
                        )
                        cargo["arrival_logged"] = True
                    break  # Only dispatch one cargo per check


    def _schedule_cargos_standard(self, now: datetime):
        """Original cargo scheduling logic (random selection)"""
        ready_count = self._count_state(READY)
        print(f"\n[SCHEDULE CHECK] Day {(now - self.start).days + 1}: {ready_count} READY tanks, First cargo scheduled: {self.first_cargo_scheduled}")

        for berth_id, berth in self.berths.items():
            if berth["current_cargo"] is None and berth["free_at"] <= now:

                random_gap_hours = random.uniform(self.berth_gap_hours_min, self.berth_gap_hours_max)
                ready_count = self._count_state(READY)
                
                # First cargo: only schedule when ready tanks are between 8-9
                if not self.first_cargo_scheduled:
                    if self.first_cargo_min_ready <= ready_count <= self.first_cargo_max_ready:
                        self.first_cargo_scheduled = True
                        arrival = now + timedelta(hours=random_gap_hours)
                    else:
                        continue  # Wait for 8-9 ready tanks for first cargo
                else:
                    # Subsequent cargos: schedule if enough tanks available
                    if ready_count < self.min_ready_tanks:
                        continue
                    
                    # Calculate arrival based on tank availability
                    time_until_empty = self._predict_next_tank_empty(now)
                    if time_until_empty is not None:
                        arrival = now + time_until_empty - timedelta(hours=18)
                        arrival = max(arrival, berth["free_at"] + timedelta(hours=random_gap_hours))
                    else:
                        arrival = berth["free_at"] + timedelta(hours=random_gap_hours)
                
                # Get all available cargo types
                available_types = [ct for ct in ["VLCC", "SUEZ", "AFRA", "PANA", "HANDY"] 
                                if ct in self.enabled_cargos and self.enabled_cargos[ct] > 0]
                
                if not available_types:
                    continue
                
                # PURE RANDOM CHOICE - NO CONSTRAINTS
                import random
                cargo_type = random.choice(available_types)
                
                # Schedule the selected cargo type
                self.cargo_counter[cargo_type] += 1
                vessel_name = f"{cargo_type}-V{self.cargo_counter[cargo_type]:03d}"
                
                volume = self.enabled_cargos[cargo_type]
                tanks_needed = self.tanks_needed_by_type[cargo_type]
                
                cargo = {
                    "vessel_name": vessel_name,
                    "cargo_type": cargo_type,
                    "berth": berth_id,
                    "arrival": arrival,
                    "fill_start": arrival + timedelta(hours=self.fill_delay_hours),
                    "volume": volume,
                    "tanks_needed": tanks_needed,
                    "tanks_started": 0,
                    "tanks_done": 0,
                    "discharge_start": None,
                    "discharge_end": None,
                    "tank_fills": [],
                    "dispatched": False
                }
                
                self.cargos.append(cargo)
                self.cargo_remaining_volume[vessel_name] = volume
                berth["current_cargo"] = cargo

                print(f"  → SCHEDULED: {vessel_name} arriving at {arrival.strftime('%d/%m/%Y %H:%M')}")
                
                self._log_event(arrival, "Success", "ARRIVAL", None, vessel_name,
                            f"BERTH {berth_id}: {vessel_name} arrives. Volume: {volume:,} bbl")

    def _maybe_finish_fill(self, now: datetime):
        """Complete fills that have reached end time"""
        finished_cargos = []
        
        for vessel_name, (tid, end_time, volume_to_fill) in list(self.active_fills.items()):
            if now >= end_time:
                # CRITICAL FIX: ADD to existing volume, don't replace it
                current_volume = self.bbl.get(tid, 0.0)
                new_volume = current_volume + volume_to_fill
                self.bbl[tid] = min(new_volume, self.usable) # self.bbl stores usable volume

                display_now_total = self.bbl[tid] + self.unusable_per_tank # This is the gross volume

                # --- START FIX ---
                # Check if the GROSS volume (display_now_total) is at full gross capacity
                total_gross_capacity = self.usable + self.unusable_per_tank
                is_tank_full = (display_now_total >= total_gross_capacity - 100)  # Allow 100 bbl tolerance
                
                event_name = "FILL_FINAL_END" if is_tank_full else "FILL_END"

                # --- START CHANGE 2 ---
                # Set state to FILLED *before* logging, if tank is full
                if is_tank_full:
                    self._change_state(tid, FILLED, end_time)
                # --- END CHANGE 2 ---
                
                # --- START FIX ---
                # Calculate remaining volume for log message
                current_cargo_remaining = self.cargo_remaining_volume.get(vessel_name, 0.0)
                remaining_after_fill = max(0.0, current_cargo_remaining - volume_to_fill)
                
                self._log_event(end_time, "Info", event_name, tid, vessel_name,
                                f"Tank {tid} fill completed: added {volume_to_fill:,.0f} bbl (now {display_now_total:,.0f} bbl). "
                                f"Cargo remaining: {remaining_after_fill:,.0f} bbl")
                # --- END FIX ---
                
                fill_start_time = end_time - timedelta(hours=volume_to_fill / self.discharge_rate)
                settle_start_time = end_time if is_tank_full else None
                settle_end_time = (end_time + timedelta(hours=self.settle_hours)) if is_tank_full else None
                lab_start_time = settle_end_time if is_tank_full and self.lab_hours > 0 else None
                ready_time_val = (end_time + timedelta(hours=self.settle_hours + self.lab_hours)) if is_tank_full else None

                # Track filling event for reports
                
                self.filling_events_log.append({
                    'tank_id': tid,
                    'start': fill_start_time.strftime('%d/%m/%Y %H:%M'),
                    'end': end_time.strftime('%d/%m/%Y %H:%M'),
                    'settle_start': end_time.strftime('%d/%m/%Y %H:%M') if is_tank_full else None,
                    'settle_end': (end_time + timedelta(hours=self.settle_hours)).strftime('%d/%m/%Y %H:%M') if is_tank_full else None,
                    'ready_time': (end_time + timedelta(hours=self.settle_hours + self.lab_hours)).strftime('%d/%m/%Y %H:%M') if is_tank_full else None
                })

                # Track discharge for daily discharge log
                cargo = next((c for c in self.cargos if c["vessel_name"] == vessel_name), None)
                if cargo:
                    self.daily_discharge_log.append({
                        'date': end_time.strftime('%d/%m/%y'),
                        'cargo_type': cargo.get('vessel_name', 'Unknown'),
                        'crude_type': cargo.get('crude_type', 'Unknown'),
                        'tank_id': tid,
                        'volume_filled': volume_to_fill
                    })
                # === END ADD ===
                
                # Handle partial or full fill
                if not is_tank_full:
                    # --- START CHANGE 1 ---
                    # Partial fill → tank goes to SUSPENDED
                    self._change_state(tid, SUSPENDED, end_time)
                    self.ready_for_fill_at[tid] = end_time + timedelta(hours=self.tank_fill_gap_hours)
                    # --- END CHANGE 1 ---
                else:
                    # --- START CHANGE 2 (Continued) ---
                    # Full fill → state is FILLED (from above), log start of settling
                    settle_end = end_time + timedelta(hours=self.settle_hours)
                    self.settle_end_at[tid] = settle_end
                    
                    # Calculate crude mix percentages
                    crude_mix_str = "Unknown"
                    if tid in self.tank_mix and self.tank_mix[tid]:
                        total_volume = sum(self.tank_mix[tid].values())
                        if total_volume > 0:
                            mix_parts = []
                            for crude, vol in self.tank_mix[tid].items():
                                pct = (vol / total_volume) * 100
                                mix_parts.append(f"{crude}: {pct:.1f}%")
                                if tid not in self.tank_mix_pct:
                                    self.tank_mix_pct[tid] = {}
                                self.tank_mix_pct[tid][crude] = pct
                            crude_mix_str = ", ".join(mix_parts)

                    # Change state from FILLED to SETTLING at end_time
                    self._change_state(tid, SETTLING, end_time)
                    
                    # Log settling start
                    self._log_event(end_time, "Info", "SETTLING_START", tid, vessel_name,
                                    f"Tank {tid} FILLED FULL ({self.bbl[tid]:,.0f} bbl) - Mix: [{crude_mix_str}] - "
                                    f"Settling for {self.settle_hours:.0f} hours")
                    # --- END CHANGE 2 (Continued) ---

                    # Handle lab testing (set timers, but DO NOT change state)
                    if self.lab_hours > 0:
                        lab_start = settle_end
                        lab_end = lab_start + timedelta(hours=self.lab_hours)
                        self.lab_start_at[tid] = lab_start # Set time for _promote_ready_tanks to check
                        self.ready_at[tid] = lab_end       # Set final ready time
                    else:
                        # No lab, tank ready immediately after settling
                        self.ready_at[tid] = settle_end

                # Update cargo progress
                cargo = next((c for c in self.cargos if c["vessel_name"] == vessel_name), None)
                if cargo:
                    cargo["tanks_done"] += 1
                    fill_start_time = end_time - timedelta(hours=volume_to_fill / self.discharge_rate)
                    cargo["tank_fills"].append((tid, fill_start_time, end_time, volume_to_fill))

                    # Update remaining volume
                    self.cargo_remaining_volume[vessel_name] -= volume_to_fill

                    # Check if cargo fully discharged
                    if self.cargo_remaining_volume[vessel_name] <= 1.0:
                        cargo["discharge_end"] = end_time
                        finished_cargos.append(cargo)

                # Remove this tank from active fills
                del self.active_fills[vessel_name]

                # Check if cargo needs to start another fill
                if cargo and self.cargo_remaining_volume[vessel_name] > 1.0:
                    
                    # Set the time when this cargo can start its *next* tank fill
                    next_available = end_time + timedelta(hours=self.tank_fill_gap_hours)
                    cargo["next_fill_available_at"] = next_available
                    
                    # Log this new gap if it's greater than 0
                    if self.tank_fill_gap_hours > 0:
                        self._log_event(end_time, "Info", "TANK_GAP_START", tid, vessel_name,
                                        f"Tank {tid} complete. {vessel_name} waiting for {self.tank_fill_gap_hours}h gap. Next fill available at {next_available.strftime('%d/%m %H:%M')}")
                    
                    self._maybe_start_fill(end_time)

        # Handle finished cargos
        for cargo in finished_cargos:
            berth = self.berths[cargo["berth"]]
            berth["current_cargo"] = None
            berth["free_at"] = cargo["discharge_end"]

            self._log_event(cargo["discharge_end"], "Success", "DISCHARGE_COMPLETE", None, cargo["vessel_name"],
                            f"BERTH {cargo['berth']}: {cargo['vessel_name']} completed discharge of {cargo['volume']:,.0f} bbl - BERTH {cargo['berth']} AVAILABLE")
            
            self.schedule_cargos(cargo["discharge_end"])
            
    def _maybe_start_fill(self, now: datetime):
        """Start new fills for arrived cargos that need tanks.
        Solver mode: use planned (tank_id, volume) partial fills.
        Standard mode: original EMPTY-tank sequential logic.
        """
        for cargo in self.cargos:
            
            # Default check: cargo must have remaining volume and not be actively filling
            base_condition = (
                self.cargo_remaining_volume.get(cargo["vessel_name"], 0) > 1.0
                and cargo["vessel_name"] not in self.active_fills
            )
            
            if not base_condition:
                continue
            
            
            # Only check fill_start time if cargo hasn't started discharging yet
            if cargo["discharge_start"] is None:
                if self.use_solver_plan:
                    if not (cargo.get('dispatched', False) == True and now >= cargo["fill_start"]):
                        continue
                else:
                    if not (now >= cargo["fill_start"]):
                        continue

            # Check if cargo is waiting for a tank fill gap (different from tank_gap_hours)
            next_fill_time = cargo.get("next_fill_available_at")
            if next_fill_time and now < next_fill_time:
                continue # Gap is not over yet, skip this cargo

            # ------------------ SOLVER-AWARE BRANCH (partial fills) ------------------
            if self.use_solver_plan and self.solver_plan_manager and hasattr(self, "cargo_to_tank_assignments"):
                cargo_key = cargo.get("cargo_id", cargo["vessel_name"])
                assigns = self.cargo_to_tank_assignments.get(cargo_key, [])

                if assigns:
                    target = None
                    for a in assigns:
                        planned_tid = a.get("tank_id")
                        
                        if isinstance(planned_tid, str):
                            match = re.match(r'TK(\d+)', planned_tid)
                            if match:
                                planned_tid = int(match.group(1))
                            else:
                                continue
                        
                        if not isinstance(planned_tid, int) or planned_tid < 1 or planned_tid > self.N:
                            continue
                        
                        planned = float(a.get("volume", 0.0))
                        filled_so_far = float(a.get("filled", 0.0))
                        assign_remaining = planned - filled_so_far

                        # --- FIX 3A: Enforce tankGapHours for solver logic ---
                        rest_time_over = now >= self.ready_for_fill_at.get(planned_tid, datetime.min)

                        if assign_remaining > 1.0 and self.state.get(planned_tid) in (EMPTY, SUSPENDED) and rest_time_over:
                        # --- END FIX 3A ---
                            current_volume = self.bbl.get(planned_tid, 0.0)
                            if current_volume < self.usable - 100:
                                target = (planned_tid, assign_remaining, a, filled_so_far, current_volume)
                                break

                    if target:
                        tid, assign_remaining, a, filled_so_far, current_volume = target
                        remaining_cargo = self.cargo_remaining_volume[cargo["vessel_name"]]
                        
                        space_in_tank = max(0, self.usable - current_volume)
                        volume_to_fill = min(assign_remaining, remaining_cargo, space_in_tank)

                        if volume_to_fill > 1.0:
                            crude_type = cargo.get("crude_type", "Unknown")
                            if tid not in self.tank_mix:
                                self.tank_mix[tid] = {}
                            if crude_type not in self.tank_mix[tid]:
                                self.tank_mix[tid][crude_type] = 0
                            self.tank_mix[tid][crude_type] += volume_to_fill
                            
                            cargo["tanks_started"] += 1
                            if cargo["discharge_start"] is None:
                                cargo["discharge_start"] = now

                            actual_fill_hours = volume_to_fill / max(self.discharge_rate, 1e-6)
                            end_time = now + timedelta(hours=actual_fill_hours)
                            self.active_fills[cargo["vessel_name"]] = (tid, end_time, volume_to_fill)

                            a["filled"] = filled_so_far + volume_to_fill

                            display_current = self.bbl[tid] + self.unusable_per_tank
                            display_target  = display_current + volume_to_fill
                            
                            event_name = "FILL_START"
                            if tid not in self.tank_filled_first:
                                event_name = "FILL_START_FIRST"
                                self.tank_filled_first.add(tid)
                            
                            vessel_name = cargo["vessel_name"]
                            if vessel_name not in self.cargo_has_started_filling:
                                message = (
                                    f"BERTH {cargo['berth']}: First fill from {vessel_name} filling Tank {tid} "
                                    f"with {volume_to_fill:,.0f} bbl {crude_type} "
                                    f"(current: {display_current:,.0f}, target: {display_target:,.0f})"
                                )
                                self.cargo_has_started_filling.add(vessel_name)
                            else:
                                message = (
                                    f"BERTH {cargo['berth']}: Start (solver) filling Tank {tid} with {volume_to_fill:,.0f} bbl {crude_type} "
                                    f"(current: {display_current:,.0f}, target: {display_target:,.0f})"
                                )
                            
                            # Change state at now
                            self._change_state(tid, FILLING, now)
                            self._log_event(now, "Info", event_name, tid, vessel_name, message)
                            
                            self.filling_events_log.append({
                                'tank_id': tid,
                                'start': now.strftime('%d/%m/%Y %H:%M'),
                                'end': None,
                                'settle_start': None,
                                'lab_start': None,
                                'ready_time': None,
                                'is_active_cycle': True
                            })

                            continue
                    continue

            # ------------------ STANDARD (existing) SEQUENTIAL LOGIC ------------------
            tid = None
            if self.initially_empty_tanks:
                # --- FIX 3B: Enforce tankGapHours for initially empty tanks ---
                tid = next((i for i in self.initially_empty_tanks 
                            if self.state[i] in (EMPTY, SUSPENDED)
                            and now >= self.ready_for_fill_at.get(i, datetime.min)), None)
                # --- END FIX 3B ---
                
                if tid:
                    self.initially_empty_tanks.remove(tid)

            if tid is None:
                # --- FIX 3C: Enforce tankGapHours for regular empty/suspended tanks ---
                # Look for EMPTY or SUSPENDED tanks that are ready for filling
                tid = next((i for i in range(1, self.N + 1)
                            if self.state[i] in (EMPTY, SUSPENDED) 
                            # NEW CHECK: Must be past the preparation time
                            and now >= self.ready_for_fill_at.get(i, datetime.min)
                            and i not in self.initially_empty_tanks), None)
                # --- END FIX 3C ---

            if tid is not None:
                remaining = self.cargo_remaining_volume[cargo["vessel_name"]]
                volume_to_fill = min(remaining, self.usable)

                cargo["tanks_started"] += 1
                if cargo["discharge_start"] is None:
                    cargo["discharge_start"] = now

                actual_fill_hours = volume_to_fill / max(self.discharge_rate, 1e-6)
                end_time = now + timedelta(hours=actual_fill_hours)
                self.active_fills[cargo["vessel_name"]] = (tid, end_time, volume_to_fill)

                event_name = "FILL_START"
                if tid not in self.tank_filled_first:
                    event_name = "FILL_START_FIRST"
                    self.tank_filled_first.add(tid)

                # Change state at now
                self._change_state(tid, FILLING, now)
                self._log_event(
                    now,
                    "Info", event_name, tid, cargo["vessel_name"],
                    f"BERTH {cargo['berth']}: Start filling Tank {tid} with {volume_to_fill:,.0f} bbl "
                    f"(rate {self.discharge_rate:,.0f} bbl/hr, duration {actual_fill_hours:.2f} h)"
                )
                
                self.filling_events_log.append({
                    'tank_id': tid,
                    'start': now.strftime('%d/%m/%Y %H:%M'),
                    'end': None,
                    'settle_start': None,
                    'lab_start': None,
                    'ready_time': None,
                    'is_active_cycle': True
                })
    
    def _find_next_ready_sequential(self, start_from: int) -> Optional[int]:
        """Find next READY tank in sequential order (1→2→3→...→N→1)"""
        # Start from next tank after current active
        for offset in range(1, self.N + 1):
            tank_id = ((start_from - 1 + offset) % self.N) + 1
            if self.state[tank_id] == READY:
                return tank_id
        return None 
    
    
    # ------------------------- FEEDING -------------------------
    def _ensure_feeding(self, now: datetime):
        """Ensure a READY tank is feeding in SEQUENTIAL ORDER"""
        # If active tank is already feeding, continue
        if self.active != 0 and self.state[self.active] == FEEDING:
            return
        
        # Find next READY tank in SEQUENTIAL ORDER starting from current active
        nxt = self._find_next_ready_sequential(self.active)
        if nxt is not None:
            # Check if we're resuming from a halt
            was_halted = self.no_feed_alert_logged
            
            self.active = nxt
            self._change_state(self.active, FEEDING, now)
            # Cap the starting volume at usable capacity to prevent overdraw
            self.bbl[self.active] = min(self.bbl[self.active], self.usable)
            self.feed_start_volume[self.active] = self.bbl[self.active]  # Track starting volume
            self.tank_feed_start_time[self.active] = now
            
            # If resuming from halt, log resume message first
            if was_halted:
                self._log_event(now, "Success", "PROCESSING_RESUME", None, None,
                              f"Processing resumed after halt")
                self.no_feed_alert_logged = False
            
            self._log_event(now, "Success", "FEED_START", self.active, None,
                          f"Tank {self.active} now starts feeding with {self.bbl[self.active]:,.0f} bbl available")
        else:
            # No READY tanks available - log only once
            if not self.no_feed_alert_logged:
                ready_count = self._count_state(READY)
                feeding_count = self._count_state(FEEDING)
                self._log_event(now, "Danger", "NO_FEED_AVAILABLE", None, None,
                              f"No tanks available for feeding. READY: {ready_count}, FEEDING: {feeding_count}")
                self.no_feed_alert_logged = True

    # scheduler.py (inside class Simulator)

    def _consume_hour(self, now: datetime, hour_end: datetime) -> float:
        """Consume for up to one hour - FIXED RATE PROCESSING"""
        # Initialize processed volume to 0.0
        processed = 0.0 
        
        # Check if we have an active feeding tank
        if self.active == 0 or self.state.get(self.active) != FEEDING:
            return processed

        if self.rate_hour <= 0:
            return processed
        
        available_in_tank =  self.bbl[self.active] 
        
        if available_in_tank <= 0:
            # Tank is empty, should not be feeding
            self._change_state(self.active, EMPTY, now)
            
            # Set ready_for_fill_at if tank is found empty unexpectedly
            self.ready_for_fill_at[self.active] = now + timedelta(hours=self.tank_gap_hours)
            
            self._log_event(now, "Warning", "FEED_ERROR", self.active, None,
                          f"Tank {self.active} marked as FEEDING but has no usable volume (current: {available_in_tank:,.0f} bbl, unusable: {self.unusable_per_tank:,.0f} bbl)")
            return processed
        
        time_to_empty_h = available_in_tank / self.rate_hour
        hour_length_h = (hour_end - now).total_seconds() / 3600.0
        
        if time_to_empty_h > hour_length_h:
            # Tank won't empty in this hour - process at FIXED RATE
            take = self.rate_hour * hour_length_h
            self.bbl[self.active] = max(0.0, self.bbl[self.active] - take)
            processed += take
            self.daily_consumption[self.active] += take
        else:
            # Tank will empty during this hour
            t_empty = now + timedelta(hours=time_to_empty_h)
            take = available_in_tank
            self.bbl[self.active] = 0
            processed += take
            emptied_tank = self.active
            self.daily_consumption[emptied_tank] += take
            
            # Calculate TOTAL draw
            total_draw = min(self.feed_start_volume[emptied_tank], self.usable)
            
            # Reset 'first_fill' flag
            if emptied_tank in self.tank_filled_first:
                self.tank_filled_first.remove(emptied_tank)
            
            # Track feeding event for reports
            if self.tank_feed_start_time.get(emptied_tank):
                self.feeding_events_log.append({
                    'tank_id': emptied_tank,
                    'start': self.tank_feed_start_time[emptied_tank].strftime('%d/%m/%Y %H:%M'),
                    'end': t_empty.strftime('%d/%m/%Y %H:%M')
                })

            # Change state at exact empty time
            self._change_state(emptied_tank, EMPTY, t_empty)
            
            # Set ready_for_fill_at using the new attribute (CRITICAL: MUST be set before logging EMPTY_START)
            self.ready_for_fill_at[emptied_tank] = t_empty + timedelta(hours=self.tank_gap_hours)
            
            # --- START FIX: Log TANK_EMPTY FIRST, then EMPTY_START ---
            
            # 1. Log TANK_EMPTY status/warning
            self._log_event(t_empty, "Warning", "TANK_EMPTY", emptied_tank, None,
                          f"Tank {emptied_tank} emptied. Total draw {total_draw:,.0f} bbl.")
            
            # 2. Log EMPTY_START (preparation time) if there is a configured gap
            if self.tank_gap_hours > 0:
                 self._log_event(t_empty, "Info", "EMPTY_START", emptied_tank, None,
                                f"Tank {emptied_tank} emptied. Preparation time of {self.tank_gap_hours}h required. Ready for fill at {self.ready_for_fill_at[emptied_tank].strftime('%d/%m %H:%M')}")
            
            # --- END FIX ---
            
            # Look for next READY tank in SEQUENTIAL ORDER
            nxt = self._find_next_ready_sequential(emptied_tank)
            if nxt is not None:
                # Check if we're resuming from a halt
                was_halted = self.no_feed_alert_logged
                
                self.active = nxt
                # Cap the volume at usable to prevent any overdraw
                self.bbl[self.active] = min(self.bbl[self.active], self.usable)
                self.feed_start_volume[self.active] = self.bbl[self.active]
                self.tank_feed_start_time[self.active] = t_empty
                
                # If resuming from halt, log resume message first
                if was_halted:
                    self._log_event(t_empty, "Success", "PROCESSING_RESUME", None, None,
                                  f"Processing resumed after halt")
                    self.no_feed_alert_logged = False
                
                # Change state at exact time
                self._change_state(self.active, FEEDING, t_empty)
                self._log_event(t_empty, "Success", "FEED_CHANGEOVER", self.active, None,
                              f"Tank {self.active} starts feeding with {self.bbl[self.active]:,.0f} bbl")
                
                # Process remainder of hour at FIXED RATE
                remaining_hour = hour_length_h - time_to_empty_h
                if remaining_hour > 0 and self.bbl[self.active] > 0:
                    additional = min(self.rate_hour * remaining_hour, self.bbl[self.active])
                    self.bbl[self.active] -= additional
                    processed += additional
                    self.daily_consumption[self.active] += additional
            else:
                # No more READY tanks - log only once
                self.active = 0
                if not self.no_feed_alert_logged:
                    ready_count = self._count_state(READY)
                    self._log_event(t_empty, "Danger", "PROCESSING_HALT", None, None,
                                  f"Processing stopped - no READY tanks available (READY: {ready_count})")
                    self.no_feed_alert_logged = True
    
        return processed

    def _promote_ready_tanks(self, now: datetime) -> int:
        """
        Promote tanks through SETTLING -> LAB -> READY states
        1. Check for SETTLING tanks that finished settling -> move to LAB (if needed) or READY
        2. Check for LAB tanks that finished testing -> move to READY
        """
        newly_ready_count = 0
        for i in range(1, self.N + 1):
            
            # --- Step 1: Check for finished SETTLING ---
            if self.state[i] == SETTLING and self.settle_end_at[i] and self.settle_end_at[i] <= now:
                settle_end_time = self.settle_end_at[i]
                
                # Case A: Tank needs lab testing
                if self.lab_hours > 0 and self.lab_start_at[i] and self.lab_start_at[i] <= now:
                    # --- Transition SETTLING -> LAB ---
                    self.settle_end_at[i] = None # Clear settle time
                    
                    lab_end_time = self.ready_at.get(i) 
                    lab_end_str = lab_end_time.strftime('%d/%m %H:%M') if lab_end_time else "Unknown"
                    
                    # Log first, then change state 1 second later
                    self._log_event(settle_end_time, "Info", "SETTLING_END", i, None,
                                    f"Settling ends. Lab testing starts for {self.lab_hours:.0f} hours (ready at {lab_end_str})")
                    self._change_state(i, LAB, settle_end_time + timedelta(seconds=1))

                # Case B: No lab testing, SETTLING -> READY
                elif self.lab_hours <= 0:
                    if self.ready_at[i] and self.ready_at[i] <= now:
                        self.bbl[i] = self.usable
                        ready_time = self.ready_at[i]
                        
                        # --- START FIX: Clear ALL timers ---
                        self.ready_at[i] = None
                        self.settle_end_at[i] = None
                        self.lab_start_at[i] = None
                        # --- END FIX ---
                        
                        newly_ready_count += 1
                        
                        self._log_event(settle_end_time, "Info", "SETTLING_END", i, None, "Settling ends")
                        
                        crude_mix_str = "Unknown"
                        if i in self.tank_mix_pct and self.tank_mix_pct[i]:
                            mix_parts = []
                            for crude, pct in self.tank_mix_pct[i].items():
                                mix_parts.append(f"{crude}: {pct:.1f}%")
                            crude_mix_str = ", ".join(mix_parts)
                        
                        # Log first, then change state 1 second later
                        self._log_event(ready_time, "Success", "READY", i, None,
                                    f"Tank {i} now READY - Mix: [{crude_mix_str}]")
                        self._change_state(i, READY, ready_time + timedelta(seconds=1))

                        if i in self.tank_cycle_counter:
                            self.tank_cycle_counter[i] += 1
                        else:
                            self.tank_cycle_counter[i] = 1
            
            # --- Step 2: Check for finished LAB (as 'elif' to prevent double-processing) ---
            elif self.state[i] == LAB and self.ready_at[i] and self.ready_at[i] <= now:
                # --- Transition LAB -> READY ---
                self.bbl[i] = self.usable
                ready_time = self.ready_at[i]
                
                # --- START FIX: Clear ALL timers ---
                self.ready_at[i] = None
                self.lab_start_at[i] = None 
                self.settle_end_at[i] = None
                # --- END FIX ---

                newly_ready_count += 1
                
                crude_mix_str = "Unknown"
                if i in self.tank_mix_pct and self.tank_mix_pct[i]:
                    mix_parts = []
                    for crude, pct in self.tank_mix_pct[i].items():
                        mix_parts.append(f"{crude}: {pct:.1f}%")
                    crude_mix_str = ", ".join(mix_parts)
                
                # Log first, then change state 1 second later
                self._log_event(ready_time, "Success", "READY", i, None,
                            f"Tank {i} now READY - Mix: [{crude_mix_str}]")
                self._change_state(i, READY, ready_time + timedelta(seconds=1))
                
                if i in self.tank_cycle_counter:
                    self.tank_cycle_counter[i] += 1
                else:
                    self.tank_cycle_counter[i] = 1
                    
        return newly_ready_count

    def _log_tank_snapshot(self, now: datetime):
        """Log complete tank inventory snapshot every 30 minutes to separate list"""
        if not hasattr(self, 'snapshot_log'):
            self.snapshot_log = []
        
        snapshot = {
            'Timestamp': now.strftime("%d/%m/%Y %H:%M"),
        }
        for i in range(1, self.N + 1):
            snapshot[f'Tank{i}'] = f"{self.bbl[i]:,.0f}"
            snapshot[f'State{i}'] = self.state[i]
        
        self.snapshot_log.append(snapshot)
    
    def simulate_day(self, day_index: int):
        day_start = self.start + timedelta(days=day_index)
        day_end = day_start + timedelta(days=1)
        simulation_end_dt = self.start + timedelta(days=self.horizon_days)
        if day_end > simulation_end_dt:
            day_end = simulation_end_dt

        # Promote SETTLING/LAB tanks to READY at day start
        newly_ready_count = self._promote_ready_tanks(day_start)

        # Reset daily consumption for all tanks
        for i in range(1, self.N + 1):
            self.daily_consumption[i] = 0.0

        # Count ready tanks at start of day
        ready_start = self._count_state(READY)
        feeding_start = self._count_state(FEEDING)
        
        # Calculate stock: tank_level - (dead_bottom + buffer_volume/2)
        ready_tanks_detail = []
        ready_stock = 0
        empty_tanks_detail = []
        for i in range(1, self.N + 1):
            if self.state[i] == READY:
                tank_usable_stock = self.bbl[i] 
                #tank_usable_stock = max(0,self.bbl[i] - self.unusable_per_tank)
                ready_stock += tank_usable_stock
                # Track tanks with non-standard stock (only for day 1 message)
                if day_index == 0:
                    ready_tanks_detail.append(f"Tank{i}: {tank_usable_stock:,.0f}")
            elif self.state[i] == EMPTY and day_index == 0:
                # Show EMPTY tanks with their current stock
                tank_empty_stock = self.bbl[i]
                empty_tanks_detail.append(f"Tank{i}: {tank_empty_stock:,.0f}")

        # FEEDING tank - only ONE at day start
        # FEEDING tanks - collect all feeding tanks with their volumes
        feeding_tanks_detail = []
        feeding_stock = 0
        for i in range(1, self.N + 1):
            if self.state[i] == FEEDING:
                tank_feed_stock = self.bbl[i]
                feeding_stock += tank_feed_stock
                feeding_tanks_detail.append(f"Tank {i}: {self.bbl[i]:,.0f} bbl")

        total_stock = ready_stock + feeding_stock
        
        # Calculate certified stock (READY + FEEDING) - this goes in DAILY_STATUS TOTAL
        certified_stock = total_stock
        
        # Calculate TRUE opening stock (ALL tanks including FILLING, SETTLING, LAB, etc.)
        true_opening_stock = sum(self.bbl[i] for i in range(1, self.N + 1))

        # Build feeding detail string
        if feeding_tanks_detail:
            feeding_detail_str = ", ".join(feeding_tanks_detail)
        else:
            feeding_detail_str = "None"

        # Build message - detailed ONLY on Day 1
        if day_index == 0:
            ready_detail_str = f" [{', '.join(ready_tanks_detail)}]" if ready_tanks_detail else ""
            empty_detail_str = f", EMPTY: [{', '.join(empty_tanks_detail)}]" if empty_tanks_detail else ""
    
            self._log_event(day_start, "Info", "DAILY_STATUS", None, None,
                f"Day starts - STOCK: READY TANKS ({ready_start}): {ready_stock:,.0f} bbl{ready_detail_str}{empty_detail_str}, FEEDING TANKS: {feeding_detail_str}, TOTAL: {certified_stock:,.0f} bbl")
        else:
            # Day 2+ or no special tanks - simple message
            self._log_event(day_start, "Info", "DAILY_STATUS", None, None,
                        f"Day starts - STOCK: READY TANKS ({ready_start}): {ready_stock:,.0f} bbl, FEEDING TANKS: {feeding_detail_str}, TOTAL: {certified_stock:,.0f} bbl")

        # Schedule cargos
        self.schedule_cargos(day_start)

        # Daily snapshot - use calculated stock
        opening_stock = true_opening_stock

       
        total_processed_today = 0.0
        
        # Calculate the precise, absolute end time for the simulation run
        simulation_end_dt = self.start + timedelta(days=self.horizon_days)
        
        # Minutely loop
        now = day_start
        snapshot_interval = timedelta(minutes=self.snapshot_interval_minutes)
        next_snapshot = day_start
        
        while now < day_end:
            
            # CRITICAL BREAK CHECK: If we're already at or past the final time, break the daily loop
            if now >= simulation_end_dt:
                 break
            
            # Log 30-minute snapshot
            if now >= next_snapshot:
                self._log_tank_snapshot(now)
                next_snapshot += snapshot_interval
                 
            # Promote any tanks that become READY during this hour
            newly_ready = self._promote_ready_tanks(now)
            
            self._maybe_finish_fill(now)
            
            self._ensure_feeding(now)
            self._maybe_start_fill(now)
            
            # Use step_end logic (use snapshot interval for step size)
            step_end = min(day_end, now + timedelta(minutes=self.snapshot_interval_minutes)) 
            
            # Cap the step_end to the absolute simulation end time
            if step_end > simulation_end_dt:
                step_end = simulation_end_dt
                
            # If the next step takes us beyond the final minute of the simulation, skip consumption.
            if now >= step_end: 
                 break 
            
            processed_this_step = self._consume_hour(now, step_end)
            total_processed_today += processed_this_step
            
            now = step_end
            
            self._maybe_finish_fill(now)
            
            # Promote tanks immediately when they become ready
            self._promote_ready_tanks(now)

        
        # --- FINAL REPORT GENERATION FIX ---
        # total_processed_today is now the correct, accumulated volume for the period run.
        final_processed_for_report = total_processed_today
        
        # Calculate Closing Stock based on accurate processing amount
        closing_stock = true_opening_stock - final_processed_for_report
        
        # Calculate TRUE closing stock (ALL tanks at end of day)
        true_closing_stock = sum(self.bbl[i] for i in range(1, self.N + 1))
        
        # Calculate opening certified stock (READY + FEEDING at day start)
        # This was already calculated at the beginning of simulate_day as 'certified_stock'
        opening_cert_stk = certified_stock  # From day start calculation
        
        # Calculate uncertified stock (opening stock - certified stock)
        opening_uncert_stk = true_opening_stock - opening_cert_stk
        
        ready_end = self._count_state(READY)
        empty_end = self._count_state(EMPTY)
        feeding_end = self._count_state(FEEDING)
        # --- END FINAL REPORT GENERATION FIX ---

        # Build feeding tanks detail - ALL tanks that fed during the day
        feeding_day_detail = []
        for i in range(1, self.N + 1):
            if self.daily_consumption[i] > 0:
                feeding_day_detail.append(f"Tank {i}: {self.daily_consumption[i]:,.0f} bbl")

        feeding_day_str = ", ".join(feeding_day_detail) if feeding_day_detail else "None"
        
        # --- START LOG MESSAGE FIX: Clean message for all days ---
        
        log_timestamp = min(now, day_end) if now < day_end else day_end - timedelta(minutes=1)
        
        # CRITICAL FIX: Force the message suffix to be empty to remove "at HH:MM hrs"
        message_suffix = "" 

        # Log end of day status
        self._log_event(log_timestamp, "Info", "DAILY_END", None, None,
                    f"Day ends{message_suffix} with {ready_end} READY tanks, FEEDING tank(s): {feeding_day_str}, Processed: {final_processed_for_report:,.0f} bbl")
        # --- END LOG MESSAGE FIX ---
        
        # Calculate certified stock (READY + FEEDING only) for inventory chart
        certified_closing_stock = sum(self.bbl[i] for i in range(1, self.N + 1) 
                                      if self.state[i] in [READY, FEEDING])
        
        # Store certified stock data for chart
        self.inventory_data.append((day_start, certified_closing_stock / 1_000_000))

        # Tank status columns
        tank_status = {f"Tank{i}": self.state[i] for i in range(1, self.N + 1)}
        
        # Final Row Population (using fixed variables)
        row = {
            "Date": day_start.strftime("%d/%m/%Y %H:%M"),
            "Opening Stock (bbl)": f"{true_opening_stock:,.0f}",
            "cert stk": f"{opening_cert_stk:,.0f}",
            "uncert stk": f"{opening_uncert_stk:,.0f}",
            "Processing (bbl)": f"{final_processed_for_report:,.0f}",
            "Closing Stock (bbl)": f"{true_closing_stock:,.0f}",
            "Ready Tanks": f"{ready_end}",
            "Empty Tanks": f"{empty_end}"
        }
        row.update(tank_status)
        self.daily_summary_rows.append(row)

    # ------------------------- RUN -------------------------
    def run(self):
        if self.infeasible:
            return

        # --- START FIX ---
        import math
        # Calculate the number of full days to loop through (e.g., if horizon is 30.5 days, loop needs 31)
        max_days = math.ceil(self.horizon_days) 

        for d in range(max_days):
            # Calculate the current day's index (0, 1, 2, ...)
            day_index = d
            
            # Stop the loop if the next day starts AFTER the total horizon time
            day_start = self.start + timedelta(days=day_index)
            
            if day_start >= self.start + timedelta(days=self.horizon_days):
                break # We've simulated all the time needed.

            self.simulate_day(day_index)
            if self.infeasible:
                break
        # --- END FIX ---
            
       
    def generate_cargo_report(self):
        """Generate cargo report with enhanced formatting"""
        
        # Clear existing rows to avoid duplicates
        self.cargo_report_rows = []
        
        # --- START NEW BLOCK (1 of 2) ---
        # 1. Pre-calculate berth gaps
        # We must sort all cargos by berth, then by arrival time to find
        # the gap between sequential cargos at the SAME berth.
        
        # Filter cargos that actually have an arrival time
        scheduled_cargos = [c for c in self.cargos if c.get('arrival')]
        
        sorted_cargos = sorted(
            scheduled_cargos,
            key=lambda c: (c['berth'], c['arrival'])
        )
        
        last_discharge_end_by_berth: Dict[int, Optional[datetime]] = {1: None, 2: None}
        berth_gaps: Dict[str, float] = {} # Maps vessel_name -> gap_hours

        for cargo in sorted_cargos:
            berth_id = cargo["berth"]
            vessel_name = cargo["vessel_name"]
            arrival_time = cargo["arrival"]
            
            # Get the discharge end time of the *previous* cargo at this *same* berth
            last_discharge_end = last_discharge_end_by_berth.get(berth_id)
            
            if last_discharge_end:
                # A previous cargo exists, calculate the gap
                gap_delta = arrival_time - last_discharge_end
                gap_hours = gap_delta.total_seconds() / 3600.0
                berth_gaps[vessel_name] = gap_hours
            else:
                # This is the first cargo at this berth, so no gap
                berth_gaps[vessel_name] = float('nan') # Use 'nan' for "N/A"

            # Update the last discharge time for this berth for the *next* iteration
            if cargo.get("discharge_end"):
                last_discharge_end_by_berth[berth_id] = cargo["discharge_end"]
        # --- END NEW BLOCK (1 of 2) ---

        # 2. Generate report rows (existing logic)
        for cargo in self.cargos:
            if cargo["discharge_start"]:
                # Calculate actual discharged volume
                actual_volume = sum(vol for _, _, _, vol in cargo["tank_fills"])
                
                # Calculate discharge duration
                discharge_duration_hours = 0
                if cargo["discharge_end"] and cargo["discharge_start"]:
                    discharge_duration_hours = (cargo["discharge_end"] - cargo["discharge_start"]).total_seconds() / 3600.0
                
                # Calculate actual tanks filled (with decimals for partial fills)
                tanks_filled = actual_volume / self.usable
                
                # Build detailed fill timestamps with separate date/time columns
                fill_details = []
                for tid, start, end, vol in cargo["tank_fills"]:
                    fill_details.append({
                        "Tank": f"Tank{tid}",
                        "Fill Start Date": start.strftime("%d/%m/%Y"),
                        "Fill Start Time": start.strftime("%H:%M"),
                        "Fill End Date": end.strftime("%d/%m/%Y"),
                        "Fill End Time": end.strftime("%H:%M"),
                        "Volume (bbl)": f"{vol:,.0f}"
                    })
                
                # --- START NEW BLOCK (2 of 2) ---
                # Get the pre-calculated gap for this cargo
                gap_hours = berth_gaps.get(cargo["vessel_name"], float('nan'))
                # Format as a string, "N/A" if it was the first cargo
                berth_gap_str = f"{gap_hours:.2f}" if not math.isnan(gap_hours) else "N/A"
                # --- END NEW BLOCK (2 of 2) ---

                row = {
                    "Vessel Name": cargo["vessel_name"],
                    "Cargo Type": cargo["cargo_type"],
                    "Berth": f"BERTH {cargo['berth']}",
                    "Arrival Date": cargo["arrival"].strftime("%d/%m/%Y"),
                    "Arrival Time": cargo["arrival"].strftime("%H:%M"),
                    "Discharge Start Date": cargo["discharge_start"].strftime("%d/%m/%Y") if cargo["discharge_start"] else "",
                    "Discharge Start Time": cargo["discharge_start"].strftime("%H:%M") if cargo["discharge_start"] else "",
                    "Discharge End Date": cargo["discharge_end"].strftime("%d/%m/%Y") if cargo["discharge_end"] else "",
                    "Discharge End Time": cargo["discharge_end"].strftime("%H:%M") if cargo["discharge_end"] else "",
                    # --- ADDED NEW COLUMN HERE ---
                    "BERTH GAP (hrs)": berth_gap_str,
                    # --- END ADD ---
                    "Discharge Duration (hours)": f"{discharge_duration_hours:.2f}",
                    "Total Volume Discharged (bbl)": f"{actual_volume:,.0f}",
                    "Tanks Filled": f"{tanks_filled:.2f}",
                    "Tank Fill Details": " | ".join([f"{d['Tank']}: {d['Tank']}: {d['Fill Start Date']} {d['Fill Start Time']}-{d['Fill End Date']} {d['Fill End Time']} ({d['Volume (bbl)']} bbl)" for d in fill_details])
                }
                self.cargo_report_rows.append(row)

    def _get_safe_filename(self, base_path):
        """Generate a safe filename if the file is already open"""
        if not os.path.exists(base_path):
            return base_path
        
        # Try to open the file to check if it's locked
        try:
            with open(base_path, 'a'):
                return base_path
        except IOError:
            # File is locked, create a new name
            directory = os.path.dirname(base_path) or "."
            filename = os.path.basename(base_path)
            name, ext = os.path.splitext(filename)
            
            counter = 1
            while True:
                new_path = os.path.join(directory, f"{name}_{counter}{ext}")
                if not os.path.exists(new_path):
                    return new_path
                try:
                    with open(new_path, 'a'):
                        return new_path
                except IOError:
                    counter += 1

    def _sort_log_chronologically(self):
        """Sort all log entries by timestamp"""
        self.daily_log_rows.sort(key=lambda x: datetime.strptime(x["Timestamp"], "%d/%m/%Y %H:%M"))

    def save_csvs(self, log_path="simulation_log.csv", 
                  summary_path="daily_summary.csv",
                  cargo_path="cargo_report.csv",
                  inventory_path="inventory_data.csv"):
        
        # --- START MODIFIED BLOCK (A): FIXED PATH SETUP ---
        # 1. Define fixed output directory relative to the project root
        output_folder = os.path.expanduser("~/Downloads")
        os.makedirs(output_folder, exist_ok=True) # Ensure the directory exists
        
        # 2. Define fixed file paths (no timestamp)
        log_path = os.path.join(output_folder, "simulation_log.csv")
        summary_path = os.path.join(output_folder, "daily_summary.csv")
        cargo_path = os.path.join(output_folder, "cargo_report.csv")
        inventory_path = os.path.join(output_folder, "inventory_data.csv")
        snapshot_path = os.path.join(output_folder, "tank_snapshots.csv")
        
        # --- END MODIFIED BLOCK (A) ---
        
        if self.infeasible:
            print(f"\n{'='*60}")
            print("SIMULATION INFEASIBLE")
            print(f"{'='*60}")
            print(f"Reason: {self.infeasible_reason}")
            print("\nThe available cargo sizes cannot maintain the required minimum READY tanks.")
            print("Suggestions:")
            print("  - Reduce minimum READY tanks requirement")
            print("  - Enable more cargo types")
            print("  - Increase number of tanks")
            print(f"{'='*60}\n")
            return
        
        # Snapshot log (uses 'w' mode, ensuring overwrite)
        if hasattr(self, 'snapshot_log') and self.snapshot_log:
            with open(snapshot_path, "w", newline="", encoding="utf-8") as f:
                fieldnames = ['Timestamp'] + [f'Tank{i}' for i in range(1, self.N + 1)] + [f'State{i}' for i in range(1, self.N + 1)]
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(self.snapshot_log)
        
        # Event log
        fieldnames = ["Timestamp", "Level", "Event", "Tank", "Cargo", "Message"]
        fieldnames += [f"Tank{i}" for i in range(1, self.N + 1)]

        self._sort_log_chronologically()
        
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.daily_log_rows)
        
        # Daily summary
        summary_fields = ["Date", "Opening Stock (bbl)", "cert stk", "uncert stk", "Processing (bbl)", 
                         "Closing Stock (bbl)", "Ready Tanks", "Empty Tanks"]
        summary_fields += [f"Tank{i}" for i in range(1, self.N + 1)]
        
        with open(summary_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=summary_fields)
            writer.writeheader()
            writer.writerows(self.daily_summary_rows)
        
        # Cargo report
        self.generate_cargo_report()
        if self.cargo_report_rows:
            with open(cargo_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=list(self.cargo_report_rows[0].keys()))
                writer.writeheader()
                writer.writerows(self.cargo_report_rows)
        


        # NOTE: The subsequent _convert_to_excel_with_autofit method will also use these fixed paths.
        
#         # Inventory data for chart
#         with open(inventory_path, "w", newline="", encoding="utf-8") as f:
#             writer = csv.writer(f)
#             writer.writerow(["Date", "Certified Stock (MMbbl)"])
#             for dt, inv in self.inventory_data:
#                 writer.writerow([dt.strftime("%d/%m/%Y"), f"{inv:.3f}"])
#         

        
       
    def _convert_to_excel_with_autofit(self, log_path, summary_path, cargo_path, inventory_path, snapshot_path):
        """Convert CSV to Excel with auto-fit columns"""
        excel_files = []
        
        for csv_path in [log_path, summary_path, cargo_path, inventory_path, snapshot_path]:
            if not os.path.exists(csv_path):
                continue
            
            excel_path = csv_path.replace('.csv', '.xlsx')
            
            # Read CSV
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            if not rows:
                continue
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write data
            for row_data in rows:
                ws.append(row_data)
            
            # Auto-fit columns with special handling for Message column
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                # Special handling for Message column (column 6)
                if col_idx == 6:  # Message column
                    adjusted_width = min(max(max_length + 10, 50), 150)  # Wider for message column
                else:
                    adjusted_width = min(max(max_length + 3, 10), 60)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Bold header row
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Save Excel
            wb.save(excel_path)
            excel_files.append(excel_path)
        
        if excel_files:
            print(f"\nSaved Excel files with auto-fit columns:")
            for ef in excel_files:
                print(f"  - {ef}")


# ------------------------- MAIN -------------------------
if __name__ == "__main__":
    print("CRUDE OIL TANK SIMULATION SYSTEM")
    
    cfg = prompt_inputs()
    sim = Simulator(cfg)
    sim.run()
    sim.save_csvs()
    
    print("Done! Check output files for detailed results.")